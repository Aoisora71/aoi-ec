"""
FastAPI server for Licel Store backend integration
Provides REST API endpoints for frontend communication
"""

from fastapi import FastAPI, HTTPException, Depends, Request, Body, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi import UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import csv
import io
import logging
import jwt
from datetime import datetime, timedelta
from typing import Optional

# Configure logging (only if not already configured)
if not logging.getLogger().handlers:
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. XLSX export/import will not work.")
from pydantic import BaseModel
from typing import List, Optional, Dict, Any, Union
import asyncio
import logging
import time
import json
import os
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Set

# Import modules
from modules.api_search import search_products, keyword_search_products, parse_keyword_search_response
from modules.db import (
    save_products_to_db,
    get_products_from_db,
    fix_products_origin_schema,
    fix_product_management_schema,
    upsert_product_management_from_origin_ids,
    get_product_management,
    get_product_management_by_item_number,
    delete_product_management_by_item_numbers,
    get_counts,
    get_recently_registered_products,
    get_category_registration_counts,
    get_product_management_stats,
    drop_removed_columns_from_products_origin,
    reset_product_management_table,
    ensure_category_management_table,
    list_categories,
    create_category_entry,
    update_category_entry,
    delete_category_entry,
    ensure_primary_category_table,
    list_primary_categories,
    create_primary_category,
    update_primary_category,
    delete_primary_category,
    save_pricing_settings,
    get_pricing_settings,
    ensure_settings_table,
    update_product_hide_item,
    update_products_hide_item_batch,
    update_all_products_hide_item,
    update_product_management_settings,
    delete_product_image,
    update_variant_selectors_with_translations,
    update_variant_selectors_and_variants,
    update_product_sku_data,
    update_single_variant,
    init_connection_pool,
    close_connection_pool,
    init_users_table,
    create_user,
    verify_user_password,
    get_user_by_email,
    get_user_by_id,
)
from modules.rakuten_product import (
    register_product_from_product_management,
    delete_product_from_product_management,
    format_error_message,
    update_product_registration_status_from_rakuten,
    update_multiple_products_registration_status_from_rakuten,
)
from modules.rakuten_inventory import register_inventory_from_product_management
from modules.enrich import enrich_products_with_detail
from modules.api_search import get_product_detail
# OpenAI-based product optimization removed

# Suppress asyncio socket errors on Windows (common when clients disconnect)
# These are harmless connection reset errors that occur when clients close connections
import sys

# Set up asyncio exception handler to suppress harmless socket errors
def asyncio_exception_handler(loop, context):
    """Handle asyncio exceptions, suppressing harmless Windows socket errors"""
    exception = context.get('exception')
    message = context.get('message', '')
    
    # Suppress common Windows socket errors that occur when clients disconnect
    if sys.platform == 'win32' and exception:
        if isinstance(exception, OSError):
            error_code = getattr(exception, 'winerror', None) or getattr(exception, 'errno', None)
            # WinError 64: The specified network name is no longer available
            if error_code == 64 or 'network name is no longer available' in str(exception).lower():
                return  # Suppress this error
            # Other common connection reset errors
            if 'connection' in str(exception).lower() and 'reset' in str(exception).lower():
                return  # Suppress connection reset errors
    
    # Suppress "Accept failed on a socket" messages
    if 'Accept failed on a socket' in message:
        return
    
    # Suppress "Task exception was never retrieved" for socket errors
    if 'Task exception was never retrieved' in message:
        if exception and isinstance(exception, OSError):
            error_code = getattr(exception, 'winerror', None) or getattr(exception, 'errno', None)
            if error_code == 64:
                return  # Suppress WinError 64 task exceptions
    
    # For other exceptions, use default handler (but suppress logging)
    # Only log if it's not a harmless socket error
    if exception and isinstance(exception, OSError):
        error_code = getattr(exception, 'winerror', None) or getattr(exception, 'errno', None)
        if error_code == 64:
            return  # Suppress all WinError 64 errors
    
    # For other exceptions, use default handler
    loop.default_exception_handler(context)

# Apply exception handler to asyncio
if sys.platform == 'win32':
    try:
        # Set exception handler for the default event loop policy
        original_policy = asyncio.get_event_loop_policy()
        
        class CustomWindowsProactorEventLoopPolicy(asyncio.WindowsProactorEventLoopPolicy):
            def new_event_loop(self):
                loop = super().new_event_loop()
                loop.set_exception_handler(asyncio_exception_handler)
                return loop
        
        asyncio.set_event_loop_policy(CustomWindowsProactorEventLoopPolicy())
        
        # Also try to set handler on existing loop if available
        try:
            loop = asyncio.get_event_loop()
            if not loop.is_running():
                loop.set_exception_handler(asyncio_exception_handler)
        except RuntimeError:
            # No event loop exists yet, which is fine
            pass
    except Exception as e:
        # If we can't set the handler, log and continue
        logger.debug(f"Could not set asyncio exception handler: {e}")

# Configure asyncio logging to suppress harmless errors
asyncio_logger = logging.getLogger('asyncio')
asyncio_logger.setLevel(logging.CRITICAL)  # Only show critical asyncio errors

# Also suppress specific Windows socket errors
if sys.platform == 'win32':
    # Filter out common Windows socket errors that are harmless
    class SocketErrorFilter(logging.Filter):
        def filter(self, record):
            # Suppress "The specified network name is no longer available" errors
            msg_str = str(getattr(record, 'msg', ''))
            message_str = str(getattr(record, 'message', ''))
            exc_info_str = str(getattr(record, 'exc_info', ''))
            
            combined = f"{msg_str} {message_str} {exc_info_str}".lower()
            
            if 'network name is no longer available' in combined:
                return False
            if 'winerror 64' in combined:
                return False
            if 'accept failed on a socket' in combined:
                return False
            if 'task exception was never retrieved' in combined:
                # Suppress task exceptions related to socket errors
                if 'winerror 64' in combined or 'network name is no longer available' in combined:
                    return False
                if 'oserror' in combined and '64' in combined:
                    return False
            return True
    
    # Apply filter to asyncio logger
    socket_filter = SocketErrorFilter()
    asyncio_logger.addFilter(socket_filter)
    
    # Also suppress errors from asyncio.proactor_events and windows_events
    proactor_logger = logging.getLogger('asyncio.proactor_events')
    proactor_logger.setLevel(logging.CRITICAL)
    proactor_logger.addFilter(socket_filter)
    
    windows_events_logger = logging.getLogger('asyncio.windows_events')
    windows_events_logger.setLevel(logging.CRITICAL)
    windows_events_logger.addFilter(socket_filter)

# Global storage for settings and logs
SETTINGS_FILE = "settings.json"
LOGS_FILE = "logs.json"
settings_data = {}
logs_data = []
refresh_task = None
refresh_keywords = set()  # Keywords to refresh automatically

def load_settings():
    """Load settings from database (with fallback to file)"""
    global settings_data
    
    # Try to load pricing settings from database first
    try:
        ensure_settings_table()
        pricing_settings = get_pricing_settings()
        
        # Load other settings from file (non-pricing settings)
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                settings_data = json.load(f)
        else:
            settings_data = {}
        
        # Merge pricing settings from database into settings_data
        shipping_costs_raw = pricing_settings.get("domestic_shipping_costs") or {}
        try:
            shipping_costs = DomesticShippingCosts(**shipping_costs_raw).dict()
        except Exception:
            base_cost = float(pricing_settings.get("domestic_shipping_cost", 300.0))
            shipping_costs = DomesticShippingCosts(
                regular=base_cost,
                size60=base_cost,
                size80=base_cost,
                size100=base_cost,
            ).dict()
        settings_data.update({
            "exchange_rate": pricing_settings.get("exchange_rate", 20.0),
            "profit_margin_percent": pricing_settings.get("profit_margin_percent", 30.0),
            "sales_commission_percent": pricing_settings.get("sales_commission_percent", 3.0),
            "currency": pricing_settings.get("currency", "JPY"),
            "domestic_shipping_cost": shipping_costs.get("regular", pricing_settings.get("domestic_shipping_cost", 300.0)),
            "domestic_shipping_costs": shipping_costs,
            "international_shipping_rate": pricing_settings.get("international_shipping_rate", 17.0),
            "customs_duty_rate": pricing_settings.get("customs_duty_rate", 100.0),
        })
        
        logger.info("✅ Loaded settings from database and file")
    except Exception as e:
        logger.warning(f"⚠️  Failed to load settings from database: {e}. Falling back to file.")
        # Fallback to file-based loading
        try:
            if os.path.exists(SETTINGS_FILE):
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    settings_data = json.load(f)
            else:
                # Default settings
                settings_data = {
                    # Pricing Settings
                    "exchange_rate": 20.0,
                    "profit_margin_percent": 30.0,
                    "sales_commission_percent": 3.0,
                    "currency": "JPY",
                    
                    # Purchase Price Calculation Settings
                    "domestic_shipping_cost": 300.0,
                    "domestic_shipping_costs": DomesticShippingCosts().dict(),
                    "international_shipping_rate": 17.0,
                    "customs_duty_rate": 100.0,
                    
                    # Server Settings
                    "auto_refresh": False,
                    "refresh_interval": 300,
                    "api_timeout": 30,
                    "max_retries": 3,
                    
                    # Logging Settings
                    "logging_enabled": True,
                    "log_level": "info",
                    "max_log_entries": 1000,
                    "log_retention_days": 30,
                    
                    # Database Settings
                    "database_url": "",
                    "connection_pool_size": 10,
                    "query_timeout": 30,
                    "enable_backup": True,
                    
                    # Login Information
                    "rakumart_api_key": "",
                    "rakumart_api_secret": "",
                    "enable_api_key_rotation": False,
                    "session_timeout": 3600,
                    
                    # User Login Information
                    "username": "",
                    "email": "",
                    "password": "",
                }
        except Exception as e2:
            logger.error(f"Failed to load settings: {e2}")
            settings_data = {}
    finally:
        # Backwards compatibility: migrate legacy fields if present
        if settings_data is not None:
            legacy_profit = settings_data.pop("profit_margin", None)
            if "profit_margin_percent" not in settings_data:
                if isinstance(legacy_profit, (int, float)):
                    settings_data["profit_margin_percent"] = legacy_profit
                else:
                    settings_data["profit_margin_percent"] = 5.0
            if "sales_commission_percent" not in settings_data:
                settings_data["sales_commission_percent"] = 10.0
            
            shipping_costs_raw = settings_data.get("domestic_shipping_costs") or {}
            base_cost = settings_data.get("domestic_shipping_cost", 300.0)
            try:
                shipping_costs_model = DomesticShippingCosts(**shipping_costs_raw)
            except Exception:
                shipping_costs_model = DomesticShippingCosts(
                    regular=base_cost,
                    size60=base_cost,
                    size80=base_cost,
                    size100=base_cost,
                )
            settings_data["domestic_shipping_costs"] = shipping_costs_model.dict()
            settings_data["domestic_shipping_cost"] = shipping_costs_model.regular

def save_settings():
    """Save settings to file"""
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings_data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logger.error(f"Failed to save settings: {e}")

def load_logs():
    """Load logs from file"""
    global logs_data
    try:
        if os.path.exists(LOGS_FILE):
            with open(LOGS_FILE, 'r', encoding='utf-8') as f:
                logs_data = json.load(f)
        else:
            logs_data = []
    except Exception as e:
        logger.error(f"Failed to load logs: {e}")
        logs_data = []

def save_logs():
    """Save logs to file"""
    try:
        with open(LOGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(logs_data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logger.error(f"Failed to save logs: {e}")

def add_log(level: str, message: str, details: str = None, source: str = "api"):
    """Add a log entry"""
    global logs_data
    if not settings_data.get("logging_enabled", True):
        return
    
    # Generate unique ID using timestamp + random component
    import random
    unique_id = f"{int(time.time() * 1000)}_{random.randint(1000, 9999)}"
    
    log_entry = {
        "id": unique_id,
        "timestamp": datetime.now().isoformat(),
        "level": level,
        "message": message,
        "details": details,
        "source": source
    }
    
    logs_data.insert(0, log_entry)  # Add to beginning
    
    # Keep only max_log_entries
    max_entries = settings_data.get("max_log_entries", 1000)
    if len(logs_data) > max_entries:
        logs_data = logs_data[:max_entries]
    
    save_logs()

async def perform_automatic_refresh():
    """Perform automatic refresh of product data"""
    global refresh_keywords
    
    if not settings_data.get("auto_refresh", False):
        return
    
    if not refresh_keywords:
        add_log("info", "No keywords configured for automatic refresh", "Add keywords to enable automatic refresh", "auto-refresh")
        return
    
    add_log("info", "Starting automatic refresh", f"Refreshing {len(refresh_keywords)} keywords", "auto-refresh")
    
    refresh_interval = settings_data.get("refresh_interval", 300)
    api_timeout = settings_data.get("api_timeout", 30)
    
    for keyword in refresh_keywords:
        try:
            add_log("info", f"Refreshing products for keyword: {keyword}", "Automatic refresh in progress", "auto-refresh")
            
            # Perform search
            products = search_products(
                keyword=keyword,
                page=1,
                page_size=50,
                shop_type="1688",
                request_timeout_seconds=api_timeout,
            )
            
            if products:
                # Save to database
                saved_count = save_products_to_db(products, keyword=keyword)
                add_log("success", f"Automatic refresh completed for '{keyword}'", f"Found {len(products)} products, saved {saved_count}", "auto-refresh")
            else:
                add_log("warning", f"No products found for keyword: {keyword}", "Automatic refresh completed with no results", "auto-refresh")
                
        except Exception as e:
            add_log("error", f"Automatic refresh failed for keyword: {keyword}", str(e), "auto-refresh")
            logger.error(f"Automatic refresh failed for keyword '{keyword}': {e}")
    
    add_log("info", "Automatic refresh cycle completed", f"Processed {len(refresh_keywords)} keywords", "auto-refresh")

async def start_auto_refresh_task():
    """Start the automatic refresh background task"""
    global refresh_task
    
    if refresh_task and not refresh_task.done():
        refresh_task.cancel()
    
    if settings_data.get("auto_refresh", False):
        refresh_interval = settings_data.get("refresh_interval", 300)
        add_log("info", "Automatic refresh task started", f"Refresh interval: {refresh_interval} seconds", "auto-refresh")
        
        async def refresh_loop():
            while True:
                try:
                    await perform_automatic_refresh()
                    await asyncio.sleep(refresh_interval)
                except asyncio.CancelledError:
                    add_log("info", "Automatic refresh task stopped", "Task was cancelled", "auto-refresh")
                    break
                except Exception as e:
                    add_log("error", "Automatic refresh task error", str(e), "auto-refresh")
                    await asyncio.sleep(refresh_interval)  # Wait before retrying
        
        refresh_task = asyncio.create_task(refresh_loop())
    else:
        add_log("info", "Automatic refresh disabled", "Auto refresh is turned off in settings", "auto-refresh")

def add_refresh_keyword(keyword: str):
    """Add a keyword to the automatic refresh list"""
    global refresh_keywords
    refresh_keywords.add(keyword)
    add_log("info", f"Added keyword to auto-refresh: {keyword}", f"Total keywords: {len(refresh_keywords)}", "auto-refresh")

def remove_refresh_keyword(keyword: str):
    """Remove a keyword from the automatic refresh list"""
    global refresh_keywords
    refresh_keywords.discard(keyword)
    add_log("info", f"Removed keyword from auto-refresh: {keyword}", f"Total keywords: {len(refresh_keywords)}", "auto-refresh")

def get_refresh_keywords():
    """Get the list of keywords for automatic refresh"""
    return list(refresh_keywords)

# Pydantic models for request/response
class ProductSearchRequest(BaseModel):
    keyword: str
    page: int = 1
    page_size: int = 50
    price_min: Optional[str] = None
    price_max: Optional[str] = None
    jpy_price_min: Optional[float] = None
    jpy_price_max: Optional[float] = None
    exchange_rate: float = 20.0
    strict_mode: bool = False
    max_length: Optional[float] = None
    max_width: Optional[float] = None
    max_height: Optional[float] = None
    max_weight: Optional[float] = None
    min_inventory: Optional[int] = None
    max_delivery_days: Optional[int] = None
    max_shipping_fee: Optional[float] = None
    shop_type: str = "1688"
    with_detail: bool = True
    detail_limit: int = 5
    save_to_db: bool = True
    categories: Optional[List[str]] = None
    subcategories: Optional[List[str]] = None

class KeywordSearchRequest(BaseModel):
    keywords: str
    shop_type: str = "1688"
    page: int = 1
    page_size: int = 50
    price_start: Optional[str] = None
    price_end: Optional[str] = None
    sort_field: Optional[str] = None
    sort_order: str = "desc"
    region_opp: Optional[str] = None
    filter: Optional[str] = None
    category_id: Optional[str] = None
    min_rating: Optional[float] = None
    min_repurchase_rate: Optional[float] = None
    min_monthly_sales: Optional[int] = None
    save_to_db: bool = True

class MultiCategorySearchRequest(BaseModel):
    category_ids: List[str]
    shop_type: str = "1688"
    page: int = 1
    page_size: int = 50
    max_products_per_category: int = 100
    price_start: Optional[str] = None
    price_end: Optional[str] = None
    sort_field: Optional[str] = None
    sort_order: Optional[str] = None
    region_opp: Optional[str] = None
    filter: Optional[str] = None
    min_rating: Optional[float] = None
    min_repurchase_rate: Optional[float] = None
    min_monthly_sales: Optional[int] = None
    save_to_db: bool = True

class BatchDeleteRequest(BaseModel):
    product_ids: List[str]


class ProductManagementBatchDeleteRequest(BaseModel):
    item_numbers: List[str]

class ProductResponse(BaseModel):
    success: bool
    data: Optional[List[Dict[str, Any]]] = None
    total: Optional[int] = None
    message: Optional[str] = None
    error: Optional[str] = None

class DatabaseResponse(BaseModel):
    success: bool
    saved_count: Optional[int] = None
    message: Optional[str] = None
    error: Optional[str] = None

class RegisterProductsRequest(BaseModel):
    product_ids: List[str]


class RegisterToRakutenRequest(BaseModel):
    item_number: str


class RegisterMultipleToRakutenRequest(BaseModel):
    item_numbers: List[str]


class RegisterInventoryToRakutenRequest(BaseModel):
    item_number: str


class RegisterMultipleInventoryToRakutenRequest(BaseModel):
    item_numbers: List[str]


class RegisterMultipleImagesToRakutenRequest(BaseModel):
    item_numbers: List[str]


class DeleteMultipleFromRakutenRequest(BaseModel):
    item_numbers: List[str]


class CheckRegistrationStatusRequest(BaseModel):
    item_number: str


class CheckMultipleRegistrationStatusRequest(BaseModel):
    item_numbers: List[str]


class DeleteImageRequest(BaseModel):
    location: str


class UpdateHideItemRequest(BaseModel):
    item_numbers: List[str]
    hide_item: bool


class DomesticShippingCosts(BaseModel):
    regular: float = 300.0
    size60: float = 360.0
    size80: float = 420.0
    size100: float = 480.0


class SettingsRequest(BaseModel):
    # Pricing Settings
    exchange_rate: float = 20.0
    profit_margin_percent: float = 5.0
    sales_commission_percent: float = 10.0
    currency: str = "JPY"
    
    # Purchase Price Calculation Settings
    domestic_shipping_cost: float = 300.0
    domestic_shipping_costs: DomesticShippingCosts = DomesticShippingCosts()
    international_shipping_rate: float = 17.0
    customs_duty_rate: float = 100.0
    
    # Server Settings
    auto_refresh: bool = False
    refresh_interval: int = 300  # seconds
    api_timeout: int = 30
    max_retries: int = 3
    
    # Logging Settings
    logging_enabled: bool = True
    log_level: str = "info"
    max_log_entries: int = 1000
    log_retention_days: int = 30
    
    # Database Settings
    database_url: Optional[str] = None
    connection_pool_size: int = 10
    query_timeout: int = 30
    enable_backup: bool = True
    
    # Login Information
    rakumart_api_key: Optional[str] = None
    rakumart_api_secret: Optional[str] = None
    enable_api_key_rotation: bool = False
    session_timeout: int = 3600
    
    # User Login Information
    username: Optional[str] = None
    email: Optional[str] = None
    password: Optional[str] = None

class LogEntry(BaseModel):
    id: str
    timestamp: str
    level: str
    message: str
    details: Optional[str] = None
    source: str

class LogsResponse(BaseModel):
    success: bool
    logs: Optional[List[LogEntry]] = None
    total_count: Optional[int] = None
    error: Optional[str] = None

class StatsResponse(BaseModel):
    success: bool
    data: Optional[Dict[str, Any]] = None
    error: Optional[str] = None

class SettingsResponse(BaseModel):
    success: bool
    settings: Optional[SettingsRequest] = None
    error: Optional[str] = None

class PrimaryCategoryRecord(BaseModel):
    id: int
    category_name: str
    default_category_ids: List[str] = []
    created_at: datetime
    updated_at: datetime


class PrimaryCategoryListResponse(BaseModel):
    success: bool
    categories: Optional[List[PrimaryCategoryRecord]] = None
    error: Optional[str] = None


class PrimaryCategoryMutationResponse(BaseModel):
    success: bool
    category: Optional[PrimaryCategoryRecord] = None
    error: Optional[str] = None


class PrimaryCategoryCreateRequest(BaseModel):
    category_name: str
    default_category_ids: Optional[List[str]] = None


class PrimaryCategoryUpdateRequest(BaseModel):
    category_name: Optional[str] = None
    default_category_ids: Optional[List[str]] = None


class CategoryAttributeGroup(BaseModel):
    name: str
    values: List[str]


class CategoryRecord(BaseModel):
    id: int
    category_name: str
    category_ids: List[str]
    rakuten_category_ids: List[str] = []
    genre_id: Optional[str] = None
    attributes: List[CategoryAttributeGroup] = []
    primary_category_id: Optional[int] = None
    primary_category_name: Optional[str] = None
    weight: Optional[float] = None
    length: Optional[float] = None
    width: Optional[float] = None
    height: Optional[float] = None
    size_option: Optional[str] = None
    size: Optional[float] = None
    created_at: datetime
    updated_at: datetime


class CategoryListResponse(BaseModel):
    success: bool
    categories: Optional[List[CategoryRecord]] = None
    error: Optional[str] = None


class CategoryMutationResponse(BaseModel):
    success: bool
    category: Optional[CategoryRecord] = None
    error: Optional[str] = None


class CategoryCreateRequest(BaseModel):
    category_name: str
    category_ids: List[str]
    rakuten_category_ids: Optional[List[str]] = None
    genre_id: Optional[str] = None
    primary_category_id: Optional[int] = None
    weight: Optional[float] = None
    length: Optional[float] = None
    width: Optional[float] = None
    height: Optional[float] = None
    size_option: Optional[str] = None
    size: Optional[float] = None
    attributes: Optional[List[CategoryAttributeGroup]] = None


class CategoryUpdateRequest(BaseModel):
    category_name: Optional[str] = None
    category_ids: Optional[List[str]] = None
    rakuten_category_ids: Optional[List[str]] = None
    genre_id: Optional[str] = None
    primary_category_id: Optional[int] = None
    weight: Optional[float] = None
    length: Optional[float] = None
    width: Optional[float] = None
    height: Optional[float] = None
    size_option: Optional[str] = None
    size: Optional[float] = None
    attributes: Optional[List[CategoryAttributeGroup]] = None

# Startup and shutdown events
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    logger.info("Starting Licel Store API server...")
    try:
        # Initialize database connection pool
        try:
            init_connection_pool(minconn=1, maxconn=20)
            logger.info("Database connection pool initialized successfully")
        except Exception as pool_error:
            logger.warning(f"Failed to initialize connection pool (server will continue with direct connections): {pool_error}")
            add_log("warning", "Connection pool initialization failed", "Server will continue but may have reduced performance. Ensure PostgreSQL is running.", "startup")
        
        # Load settings and logs
        load_settings()
        load_logs()
        logger.info("Settings and logs loaded successfully")
        
        # Initialize database schema if needed (gracefully handle connection errors)
        try:
            fix_products_origin_schema()
            fix_product_management_schema()
            ensure_primary_category_table()
            ensure_category_management_table()
            init_users_table()  # Initialize users table for authentication
            logger.info("Database schema initialized successfully")
        except Exception as db_error:
            logger.warning(f"Database connection failed during startup (server will continue): {db_error}")
            add_log("warning", "Database not available during startup", "Server will continue but database features may be unavailable. Ensure PostgreSQL is running.", "startup")
        
        # Add startup log
        add_log("info", "Licel Store API server started", "Server initialization completed", "startup")
        
        # Start automatic refresh task if enabled
        await start_auto_refresh_task()
    except Exception as e:
        logger.error(f"Failed to initialize server: {e}")
        add_log("error", "Server initialization failed", str(e), "startup")
    
    yield
    
    # Shutdown
    logger.info("Shutting down Licel Store API server...")
    
    # Stop automatic refresh task
    global refresh_task
    if refresh_task and not refresh_task.done():
        refresh_task.cancel()
        try:
            await refresh_task
        except asyncio.CancelledError:
            pass
    
    # Close database connection pool
    try:
        close_connection_pool()
        logger.info("Database connection pool closed successfully")
    except Exception as pool_error:
        logger.warning(f"Error closing connection pool: {pool_error}")
    
    add_log("info", "Licel Store API server shutting down", "Server shutdown initiated", "shutdown")

# Create FastAPI app
app = FastAPI(
    title="Licel Store API",
    description="REST API for Licel Store product management system",
    version="1.0.0",
    lifespan=lifespan
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for development
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Add middleware to log 404 requests with details

@app.middleware("http")
async def log_404_requests(request: Request, call_next):
    response = await call_next(request)
    if response.status_code == 404:
        client_ip = request.client.host if request.client else "unknown"
        user_agent = request.headers.get("user-agent", "unknown")
        referer = request.headers.get("referer", "none")
        
        logger.warning(
            f"404 Not Found: {request.method} {request.url.path} | "
            f"IP: {client_ip} | "
            f"User-Agent: {user_agent} | "
            f"Referer: {referer}"
        )
        
        add_log(
            "warning",
            f"404 Not Found: {request.method} {request.url.path}",
            f"IP: {client_ip}, User-Agent: {user_agent}, Referer: {referer}",
            "api"
        )
    
    return response

# ============================================================================
# JWT Token Management
# ============================================================================

# JWT Configuration
JWT_SECRET_KEY = os.getenv("JWT_SECRET_KEY", "your-secret-key-change-in-production")
JWT_ALGORITHM = "HS256"
JWT_ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24  # 24 hours
JWT_REFRESH_TOKEN_EXPIRE_DAYS = 7  # 7 days

security = HTTPBearer()

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    """Create a JWT access token."""
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=JWT_ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire, "type": "access"})
    encoded_jwt = jwt.encode(to_encode, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)
    return encoded_jwt

def create_refresh_token(data: dict) -> str:
    """Create a JWT refresh token."""
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(days=JWT_REFRESH_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire, "type": "refresh"})
    encoded_jwt = jwt.encode(to_encode, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)
    return encoded_jwt

def verify_token(token: str) -> Optional[dict]:
    """Verify and decode a JWT token."""
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        logger.warning("Token has expired")
        return None
    except jwt.InvalidTokenError as e:
        logger.warning(f"Invalid token: {e}")
        return None

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    """Dependency to get current authenticated user from JWT token."""
    token = credentials.credentials
    payload = verify_token(token)
    if payload is None:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    if payload.get("type") != "access":
        raise HTTPException(status_code=401, detail="Invalid token type")
    
    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token payload")
    
    user = get_user_by_id(user_id)
    if not user or not user.get("is_active"):
        raise HTTPException(status_code=401, detail="User not found or inactive")
    
    return user

# ============================================================================
# Authentication Endpoints
# ============================================================================

class LoginRequest(BaseModel):
    email: str
    password: str

class SignupRequest(BaseModel):
    email: str
    password: str
    name: str

class TokenResponse(BaseModel):
    access_token: str
    refresh_token: str
    token_type: str = "bearer"
    user: dict

@app.post("/api/auth/login", response_model=TokenResponse)
async def login(request: LoginRequest):
    """Authenticate user and return JWT tokens."""
    try:
        # Validate input
        if not request.email or not request.email.strip():
            raise HTTPException(status_code=400, detail="メールアドレスを入力してください")
        if not request.password:
            raise HTTPException(status_code=400, detail="パスワードを入力してください")
        
        # Ensure users table exists
        try:
            init_users_table()
        except Exception as table_error:
            logger.warning(f"Users table initialization warning: {table_error}")
        
        # Verify credentials
        try:
            user = verify_user_password(request.email.strip(), request.password)
        except Exception as verify_error:
            error_msg = str(verify_error)
            logger.error(f"Error verifying user credentials: {error_msg}", exc_info=True)
            if "DSN is not configured" in error_msg:
                raise HTTPException(status_code=500, detail="データベース接続が設定されていません。管理者にお問い合わせください。")
            elif "relation \"users\" does not exist" in error_msg.lower():
                # Table doesn't exist, try to create it
                try:
                    init_users_table()
                    user = verify_user_password(request.email.strip(), request.password)
                except Exception as retry_error:
                    logger.error(f"Failed to create users table: {retry_error}")
                    raise HTTPException(status_code=500, detail="データベーステーブルの作成に失敗しました。管理者にお問い合わせください。")
            else:
                raise HTTPException(status_code=500, detail=f"ログイン処理中にエラーが発生しました: {error_msg}")
        
        if not user:
            logger.warning(f"Failed login attempt for email: {request.email}")
            raise HTTPException(status_code=401, detail="メールアドレスまたはパスワードが正しくありません")
        
        # Create tokens
        try:
            access_token = create_access_token(data={"sub": user["id"], "email": user["email"]})
            refresh_token = create_refresh_token(data={"sub": user["id"], "email": user["email"]})
        except Exception as token_error:
            logger.error(f"Token creation error: {token_error}")
            raise HTTPException(status_code=500, detail="トークンの生成に失敗しました。もう一度お試しください。")
        
        logger.info(f"User {user['email']} logged in successfully")
        
        return TokenResponse(
            access_token=access_token,
            refresh_token=refresh_token,
            user=user
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"ログイン中にエラーが発生しました: {str(e)}")

@app.post("/api/auth/signup", response_model=TokenResponse)
async def signup(request: SignupRequest):
    """Create a new user account and return JWT tokens."""
    try:
        # Validate input
        if not request.email or not request.email.strip():
            raise HTTPException(status_code=400, detail="メールアドレスを入力してください")
        if not request.password:
            raise HTTPException(status_code=400, detail="パスワードを入力してください")
        if not request.name or not request.name.strip():
            raise HTTPException(status_code=400, detail="名前を入力してください")
        
        # Validate email format
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, request.email.strip()):
            raise HTTPException(status_code=400, detail="有効なメールアドレスを入力してください")
        
        # Validate password length
        if len(request.password) < 6:
            raise HTTPException(status_code=400, detail="パスワードは6文字以上である必要があります")
        
        # Ensure users table exists
        try:
            init_users_table()
            logger.info("Users table initialized/verified")
        except Exception as table_error:
            error_msg = str(table_error)
            if "already exists" not in error_msg.lower() and "duplicate" not in error_msg.lower():
                logger.warning(f"Users table initialization warning: {table_error}")
        
        # Create user
        user = None
        try:
            logger.info(f"Attempting to create user: {request.email.strip()}")
            user = create_user(request.email.strip(), request.password, request.name.strip())
            if not user:
                logger.warning(f"User creation returned None for email: {request.email.strip()}")
                raise HTTPException(status_code=400, detail="このメールアドレスは既に登録されています")
            logger.info(f"User created successfully: {user.get('email')} (ID: {user.get('id')})")
        except ValueError as ve:
            logger.warning(f"Validation error during signup: {ve}")
            raise HTTPException(status_code=400, detail=str(ve))
        except RuntimeError as re:
            # Database connection error
            error_msg = str(re)
            logger.error(f"Database error during signup: {error_msg}", exc_info=True)
            if "DSN is not configured" in error_msg:
                raise HTTPException(status_code=500, detail="データベース接続が設定されていません。管理者にお問い合わせください。")
            elif "connection" in error_msg.lower() or "timeout" in error_msg.lower():
                raise HTTPException(status_code=500, detail="データベースに接続できません。しばらく待ってから再試行してください。")
            elif "relation \"users\" does not exist" in error_msg.lower():
                # Table doesn't exist, try to create it
                try:
                    logger.info("Users table not found, attempting to create...")
                    init_users_table()
                    # Retry user creation
                    user = create_user(request.email.strip(), request.password, request.name.strip())
                    if not user:
                        raise HTTPException(status_code=400, detail="このメールアドレスは既に登録されています")
                    logger.info(f"User created after table creation: {user.get('email')}")
                except Exception as retry_error:
                    logger.error(f"Failed to create users table: {retry_error}", exc_info=True)
                    raise HTTPException(status_code=500, detail="データベーステーブルの作成に失敗しました。管理者にお問い合わせください。")
            else:
                raise HTTPException(status_code=500, detail=f"データベースエラー: {error_msg}")
        except Exception as db_error:
            # Other database errors
            error_msg = str(db_error)
            logger.error(f"Database error during user creation: {error_msg}", exc_info=True)
            if "duplicate key" in error_msg.lower() or "unique constraint" in error_msg.lower() or "already exists" in error_msg.lower():
                raise HTTPException(status_code=400, detail="このメールアドレスは既に登録されています")
            elif "relation \"users\" does not exist" in error_msg.lower():
                # Table doesn't exist, try to create it
                try:
                    logger.info("Users table not found in exception handler, attempting to create...")
                    init_users_table()
                    # Retry user creation
                    user = create_user(request.email.strip(), request.password, request.name.strip())
                    if not user:
                        raise HTTPException(status_code=400, detail="このメールアドレスは既に登録されています")
                    logger.info(f"User created after table creation in exception handler: {user.get('email')}")
                except Exception as retry_error:
                    logger.error(f"Failed to create users table in exception handler: {retry_error}", exc_info=True)
                    raise HTTPException(status_code=500, detail="データベーステーブルの作成に失敗しました。管理者にお問い合わせください。")
            else:
                raise HTTPException(status_code=500, detail=f"アカウント作成中にエラーが発生しました: {error_msg}")
        
        # Verify user was created
        if not user:
            logger.error("User creation failed: user is None")
            raise HTTPException(status_code=500, detail="ユーザーの作成に失敗しました。もう一度お試しください。")
        
        # Create tokens
        try:
            access_token = create_access_token(data={"sub": user["id"], "email": user["email"]})
            refresh_token = create_refresh_token(data={"sub": user["id"], "email": user["email"]})
        except Exception as token_error:
            logger.error(f"Token creation error: {token_error}")
            raise HTTPException(status_code=500, detail="トークンの生成に失敗しました。もう一度お試しください。")
        
        logger.info(f"New user {user['email']} signed up successfully")
        
        return TokenResponse(
            access_token=access_token,
            refresh_token=refresh_token,
            user={
                "id": str(user["id"]),
                "email": user["email"],
                "name": user["name"],
                "is_active": user["is_active"]
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Signup error: {e}", exc_info=True)
        error_msg = str(e)
        # Don't expose internal errors, but log them
        if "detail" in error_msg.lower() or "HTTPException" in error_msg:
            raise HTTPException(status_code=500, detail="アカウント作成中にエラーが発生しました。もう一度お試しください。")
        raise HTTPException(status_code=500, detail=f"アカウント作成中にエラーが発生しました: {error_msg}")

@app.post("/api/auth/verify")
async def verify_token_endpoint(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Verify JWT token and return user data."""
    try:
        user = await get_current_user(credentials)
        return {"valid": True, "user": user}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Token verification error: {e}")
        raise HTTPException(status_code=401, detail="Invalid token")

@app.post("/api/auth/refresh")
async def refresh_token_endpoint(refresh_token: str = Body(..., embed=True)):
    """Refresh access token using refresh token."""
    try:
        payload = verify_token(refresh_token)
        if not payload or payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid refresh token")
        
        user_id = payload.get("sub")
        user = get_user_by_id(user_id)
        if not user or not user.get("is_active"):
            raise HTTPException(status_code=401, detail="User not found or inactive")
        
        # Create new access token
        access_token = create_access_token(data={"sub": user["id"], "email": user["email"]})
        
        return {
            "access_token": access_token,
            "token_type": "bearer"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Token refresh error: {e}")
        raise HTTPException(status_code=401, detail="Invalid refresh token")

@app.get("/api/auth/me")
async def get_current_user_info(current_user: dict = Depends(get_current_user)):
    """Get current authenticated user information."""
    return {"user": current_user}

@app.get("/")
async def root():
    return {"message": "Licel Store API Server", "status": "running"}

@app.get("/api/health")
async def api_health():
    """Health check endpoint for API connectivity testing."""
    return {
        "status": "healthy",
        "service": "licel-store-api",
        "timestamp": datetime.utcnow().isoformat()
    }

@app.get("/health")
async def health_check():
    return {"status": "healthy", "service": "licel-store-api"}

@app.get("/health/full")
async def full_health_check():
    """
    Return comprehensive service status for dashboard.
    """
    status: Dict[str, Any] = {
        "service": "licel-store-api",
        "status": "healthy",
        "uptime_ok": True,
        "settings": {
            "loaded": bool(settings_data),
            "auto_refresh": settings_data.get("auto_refresh", False),
            "refresh_interval": settings_data.get("refresh_interval", 300),
            "logging_enabled": settings_data.get("logging_enabled", True),
        },
        "subsystems": {
            "database": {"ok": False, "details": None},
            "auto_refresh": {"ok": False, "running": False},
            "rakumart_config": {"ok": False},
            "logs_store": {"ok": False},
        },
    }

    # Database check
    try:
        _ = get_products_from_db(limit=1)
        status["subsystems"]["database"] = {"ok": True, "details": "reachable"}
    except Exception as e:
        status["subsystems"]["database"] = {"ok": False, "details": str(e)}
        status["status"] = "degraded"

    # Auto refresh task status
    try:
        running = bool(refresh_task and not refresh_task.done())
        status["subsystems"]["auto_refresh"] = {
            "ok": True,
            "running": running,
            "keywords_count": len(refresh_keywords),
        }
    except Exception as e:
        status["subsystems"]["auto_refresh"] = {"ok": False, "details": str(e)}
        status["status"] = "degraded"

    # Rakumart config presence (basic sanity)
    try:
        has_keys = bool(settings_data.get("rakumart_api_key") is not None)
        status["subsystems"]["rakumart_config"] = {"ok": True, "has_api_key": has_keys}
    except Exception:
        status["subsystems"]["rakumart_config"] = {"ok": False}
        status["status"] = "degraded"

    # Logs store writeability
    try:
        # attempt a no-op save
        save_logs()
        status["subsystems"]["logs_store"] = {"ok": True}
    except Exception as e:
        status["subsystems"]["logs_store"] = {"ok": False, "details": str(e)}
        status["status"] = "degraded"

    return status

@app.options("/api/products/search")
async def options_search():
    return {}

@app.options("/api/products/{path:path}")
async def options_products():
    return {}

@app.get("/api/test")
async def test_endpoint():
    """Test endpoint to verify connectivity"""
    return {"message": "Test endpoint working", "status": "ok", "timestamp": str(time.time())}

@app.post("/api/products/search", response_model=ProductResponse)
async def search_products_endpoint(request: ProductSearchRequest):
    """
    Search products using the legacy search API
    """
    try:
        logger.info(f"Product search request: {request.keyword}")
        add_log("info", f"Product search started", f"Keyword: {request.keyword}, Page: {request.page}, Page size: {request.page_size}", "search")
        
        # Perform search with all filters
        products = search_products(
            keyword=request.keyword,
            page=request.page,
            page_size=request.page_size,
            price_min=request.price_min,
            price_max=request.price_max,
            jpy_price_min=request.jpy_price_min,
            jpy_price_max=request.jpy_price_max,
            exchange_rate=request.exchange_rate,
            strict_mode=request.strict_mode,
            max_length=request.max_length,
            max_width=request.max_width,
            max_height=request.max_height,
            max_weight=request.max_weight,
            min_inventory=request.min_inventory,
            max_delivery_days=request.max_delivery_days,
            max_shipping_fee=request.max_shipping_fee,
            categories=request.categories,
            subcategories=request.subcategories,
            shop_type=request.shop_type,
            request_timeout_seconds=30,
        )
        
        # Enrich with details if requested (fetch full detail)
        if request.with_detail and products:
            from modules.api_search import get_product_detail
            def _fetch_detail(**kwargs):
                try:
                    return get_product_detail(
                        goods_id=kwargs.get("goods_id"),
                        shop_type=request.shop_type,
                        request_timeout_seconds=request.request_timeout_seconds if hasattr(request, 'request_timeout_seconds') else 30,
                        api_url=None,
                        normalize=True,
                    )
                except Exception:
                    return None
            enrich_products_with_detail(
                products,
                get_detail_fn=_fetch_detail,
                shop_type=request.shop_type,
                request_timeout_seconds=30,
                limit=request.detail_limit,
            )
        
        # Save to database if requested
        saved_count = 0
        if request.save_to_db and products:
            try:
                saved_count = save_products_to_db(products, keyword=request.keyword)
                logger.info(f"Saved {saved_count} products to database")
            except Exception as e:
                logger.error(f"Failed to save products to database: {e}")
                # Try to fix schema and retry
                try:
                    fix_products_origin_schema()
                    saved_count = save_products_to_db(products, keyword=request.keyword)
                    logger.info(f"Fixed schema and saved {saved_count} products to database")
                except Exception as e2:
                    logger.error(f"Failed to save products even after schema fix: {e2}")
        
        add_log("success", f"Product search completed successfully", f"Found {len(products)} products, saved {saved_count} to database", "search")
        
        return ProductResponse(
            success=True,
            data=products,
            total=len(products),
            message=f"Found {len(products)} products, saved {saved_count} to database"
        )
        
    except Exception as e:
        logger.error(f"Product search failed: {e}")
        add_log("error", "Product search failed", str(e), "search")
        return ProductResponse(
            success=False,
            error=str(e),
            message="Product search failed"
        )

@app.post("/api/products/keyword-search", response_model=ProductResponse)
async def keyword_search_endpoint(request: KeywordSearchRequest):
    """
    Search products using the new keyword search API
    """
    try:
        logger.info(f"Keyword search request: keywords='{request.keywords}', shop_type='{request.shop_type}', page={request.page}, page_size={request.page_size}")
        logger.info(f"Filter parameters: price_start='{request.price_start}', price_end='{request.price_end}', region_opp='{request.region_opp}', filter='{request.filter}', category_id='{request.category_id}'")
        
        # Validate and log price parameters
        if request.price_start:
            try:
                price_start_float = float(request.price_start)
                logger.info(f"Price start validation: '{request.price_start}' -> {price_start_float}")
            except ValueError:
                logger.warning(f"Invalid price_start format: '{request.price_start}' - not a valid number")
        
        if request.price_end:
            try:
                price_end_float = float(request.price_end)
                logger.info(f"Price end validation: '{request.price_end}' -> {price_end_float}")
            except ValueError:
                logger.warning(f"Invalid price_end format: '{request.price_end}' - not a valid number")
        
        # Validate price range
        if request.price_start and request.price_end:
            try:
                start_val = float(request.price_start)
                end_val = float(request.price_end)
                if start_val > end_val:
                    logger.warning(f"Invalid price range: start ({start_val}) > end ({end_val})")
                else:
                    logger.info(f"Valid price range: {start_val} - {end_val}")
            except ValueError:
                logger.warning("Could not validate price range due to invalid number format")
        
        # Validate required parameters
        if not request.keywords or request.keywords.strip() == "":
            logger.error("Empty keywords provided")
            return ProductResponse(
                success=False,
                error="Keywords are required",
                message="Please provide search keywords"
            )
        
        # Prepare sort parameter
        sort_param = None
        if request.sort_field:
            sort_param = {request.sort_field: request.sort_order}
            logger.info(f"Sort parameter: {sort_param}")
        
        # Perform keyword search
        logger.info("Calling keyword_search_products...")
        response = keyword_search_products(
            keywords=request.keywords,
            shop_type=request.shop_type,
            page=request.page,
            page_size=request.page_size,
            price_start=request.price_start,
            price_end=request.price_end,
            sort=sort_param,
            region_opp=request.region_opp,
            filter=request.filter,
            category_id=request.category_id,
            request_timeout_seconds=30,
        )
        
        logger.info(f"Raw API response received: {type(response)}")
        
        # Parse response
        logger.info("Parsing API response...")
        parsed_response = parse_keyword_search_response(response)
        
        if not parsed_response.get("success", False):
            error_msg = parsed_response.get("error", "Unknown error")
            logger.error(f"API search failed: {error_msg}")
            add_log("error", "Product search failed", f"Keywords: {request.keywords}, Error: {error_msg}", "search")
            return ProductResponse(
                success=False,
                error=f"API search failed: {error_msg}",
                message="Failed to search products"
            )
        
        products = parsed_response.get("data", {}).get("result", {}).get("products", [])
        total = parsed_response.get("data", {}).get("result", {}).get("total", 0)
        
        # Apply quality filters if specified
        original_count = len(products)
        if products and (request.min_rating is not None or request.min_repurchase_rate is not None or request.min_monthly_sales is not None):
            filtered_products = []
            for product in products:
                # Filter by minimum rating
                if request.min_rating is not None:
                    rating = product.get("tradeScore") or product.get("rating") or product.get("trade_score") or 0
                    try:
                        rating_float = float(rating) if rating else 0
                        if rating_float < request.min_rating:
                            continue
                    except (ValueError, TypeError):
                        continue
                
                # Filter by minimum repurchase rate
                if request.min_repurchase_rate is not None:
                    repurchase = product.get("repurchaseRate") or product.get("rePurchaseRate") or product.get("repurchase_rate") or 0
                    try:
                        repurchase_float = float(repurchase) if repurchase else 0
                        if repurchase_float < request.min_repurchase_rate:
                            continue
                    except (ValueError, TypeError):
                        continue
                
                # Filter by minimum monthly sales
                if request.min_monthly_sales is not None:
                    sales = product.get("monthSold") or product.get("month_sold") or product.get("monthSoldCount") or 0
                    try:
                        sales_int = int(sales) if sales else 0
                        if sales_int < request.min_monthly_sales:
                            continue
                    except (ValueError, TypeError):
                        continue
                
                filtered_products.append(product)
            
            products = filtered_products
            logger.info(f"Applied quality filters: {len(products)}/{original_count} products remaining")
            if request.min_rating is not None:
                logger.info(f"  - Min rating: {request.min_rating}")
            if request.min_repurchase_rate is not None:
                logger.info(f"  - Min repurchase rate: {request.min_repurchase_rate}%")
            if request.min_monthly_sales is not None:
                logger.info(f"  - Min monthly sales: {request.min_monthly_sales}")
        
        # Post-process price filtering to ensure accuracy
        if products and (request.price_start or request.price_end):
            logger.info(f"Post-processing price filter: start={request.price_start}, end={request.price_end}")
            original_count = len(products)
            
            try:
                price_start_val = float(request.price_start) if request.price_start else None
                price_end_val = float(request.price_end) if request.price_end else None
                
                filtered_products = []
                for product in products:
                    product_price = product.get("goodsPrice")
                    if product_price is None:
                        logger.warning(f"Product {product.get('goodsId')} has no price, skipping price filter")
                        filtered_products.append(product)
                        continue
                    
                    try:
                        price_val = float(product_price)
                        
                        # Apply price filters
                        if price_start_val is not None and price_val < price_start_val:
                            logger.debug(f"Product {product.get('goodsId')} price {price_val} below start {price_start_val}")
                            continue
                        if price_end_val is not None and price_val > price_end_val:
                            logger.debug(f"Product {product.get('goodsId')} price {price_val} above end {price_end_val}")
                            continue
                        
                        filtered_products.append(product)
                        logger.debug(f"Product {product.get('goodsId')} price {price_val} within range")
                        
                    except (ValueError, TypeError):
                        logger.warning(f"Product {product.get('goodsId')} has invalid price format: {product_price}")
                        filtered_products.append(product)  # Include products with invalid prices
                
                products = filtered_products
                logger.info(f"Price filtering: {original_count} -> {len(products)} products")
                
            except ValueError as e:
                logger.error(f"Error in price filtering: {e}")
                # Keep original products if filtering fails
        
        # Enrich a subset of products with detail payload so DB has rich fields
        try:
            if products:
                enrich_limit = min(len(products), 50)
                enrich_products_with_detail(
                    products,
                    get_detail_fn=lambda **kwargs: get_product_detail(
                        goods_id=kwargs.get("goods_id"),
                        shop_type=kwargs.get("shop_type", request.shop_type),
                        request_timeout_seconds=30,
                        normalize=True,
                    ),
                    shop_type=request.shop_type,
                    request_timeout_seconds=30,
                    limit=enrich_limit,
                )
        except Exception as e:
            logger.warning(f"Enrich step failed: {e}")

        # If no products found with keyword search, try fallback search
        if not products and request.keywords:
            logger.info("No products found with keyword search, trying fallback search...")
            try:
                from modules.api_search import search_products
                fallback_products = search_products(
                    keyword=request.keywords,
                    page=request.page,
                    page_size=request.page_size,
                    shop_type=request.shop_type,
                    price_min=request.price_start,
                    price_max=request.price_end,
                    order_key=request.sort_field,
                    order_value=request.sort_order,
                    request_timeout_seconds=30,
                )
                
                if fallback_products:
                    logger.info(f"Fallback search found {len(fallback_products)} products")
                    # Convert fallback products to the expected format
                    products = []
                    for p in fallback_products:
                        converted_product = {
                            "shopType": p.get("shop_type"),
                            "goodsId": p.get("goods_id"),
                            "titleC": p.get("title_c"),
                            "titleT": p.get("title_t"),
                            "traceInfo": p.get("trace_info"),
                            "goodsPrice": p.get("goods_price"),
                            "imgUrl": p.get("img_url"),
                            "monthSold": p.get("month_sold"),
                            "isJxhy": p.get("is_jxhy"),
                            "goodsTags": p.get("goods_tags"),
                            "shopCity": p.get("shop_city"),
                            "repurchaseRate": p.get("repurchase_rate"),
                            "sellerIdentities": p.get("seller_identities"),
                        }
                        products.append(converted_product)
                    
                    # Apply price filtering to fallback results as well
                    if products and (request.price_start or request.price_end):
                        logger.info(f"Applying price filter to fallback results: start={request.price_start}, end={request.price_end}")
                        original_count = len(products)
                        
                        try:
                            price_start_val = float(request.price_start) if request.price_start else None
                            price_end_val = float(request.price_end) if request.price_end else None
                            
                            filtered_products = []
                            for product in products:
                                product_price = product.get("goodsPrice")
                                if product_price is None:
                                    filtered_products.append(product)
                                    continue
                                
                                try:
                                    price_val = float(product_price)
                                    
                                    # Apply price filters
                                    if price_start_val is not None and price_val < price_start_val:
                                        continue
                                    if price_end_val is not None and price_val > price_end_val:
                                        continue
                                    
                                    filtered_products.append(product)
                                    
                                except (ValueError, TypeError):
                                    filtered_products.append(product)  # Include products with invalid prices
                            
                            products = filtered_products
                            logger.info(f"Fallback price filtering: {original_count} -> {len(products)} products")
                            
                        except ValueError as e:
                            logger.error(f"Error in fallback price filtering: {e}")
                    
                    total = len(products)
                    add_log("success", "Fallback search successful", f"Keywords: {request.keywords}, Found: {len(products)}", "search")
                else:
                    logger.warning("Fallback search also returned no products")
                    add_log("warning", "No products found in fallback search", f"Keywords: {request.keywords}", "search")
                    
            except Exception as e:
                logger.error(f"Fallback search failed: {e}")
                add_log("error", "Fallback search failed", f"Keywords: {request.keywords}, Error: {str(e)}", "search")
        
        logger.info(f"Final result: {len(products)} products (total: {total})")
        
        # Save to database if requested
        saved_count = 0
        if request.save_to_db and products:
            try:
                logger.info("Saving products to database...")
                saved_count = save_products_to_db(products, keyword=request.keywords)
                logger.info(f"Saved {saved_count} products to database")
                add_log("success", f"Products saved to database", f"Keywords: {request.keywords}, Saved: {saved_count}/{len(products)}", "database")
            except Exception as e:
                logger.error(f"Failed to save products to database: {e}")
                try:
                    fix_products_origin_schema()
                    saved_count = save_products_to_db(products, keyword=request.keywords)
                    logger.info(f"Fixed schema and saved {saved_count} products to database")
                    add_log("success", f"Schema fixed and products saved", f"Keywords: {request.keywords}, Saved: {saved_count}/{len(products)}", "database")
                except Exception as e2:
                    logger.error(f"Failed to save products even after schema fix: {e2}")
                    add_log("error", "Database save failed", f"Keywords: {request.keywords}, Error: {e2}", "database")
        
        add_log("success", f"Product search completed", f"Keywords: {request.keywords}, Found: {len(products)}, Saved: {saved_count}", "search")
        
        return ProductResponse(
            success=True,
            data=products,
            total=total,
            message=f"Found {len(products)} products, saved {saved_count} to database"
        )
        
    except Exception as e:
        logger.error(f"Keyword search failed: {e}")
        add_log("error", "Keyword search failed", f"Keywords: {request.keywords}, Error: {str(e)}", "search")
        return ProductResponse(
            success=False,
            error=str(e),
            message="Keyword search failed"
        )

@app.post("/api/products/multi-category-search", response_model=ProductResponse)
async def multi_category_search_endpoint(request: MultiCategorySearchRequest):
    """
    Search products across multiple categories
    """
    try:
        logger.info(f"Multi-category search request received: {len(request.category_ids)} categories")
        logger.info(f"Category IDs: {request.category_ids}")
        logger.info(f"Request parameters: shop_type={request.shop_type}, page={request.page}, page_size={request.page_size}, max_products_per_category={request.max_products_per_category}")
        
        # Validate request
        if not request.category_ids or len(request.category_ids) == 0:
            logger.error("No category IDs provided in request")
            return ProductResponse(
                success=False,
                error="No category IDs provided",
                message="At least one category ID is required"
            )
        
        from modules.api_search import search_multiple_categories
        
        # Prepare sort parameter
        sort_param = None
        if request.sort_field:
            sort_param = {request.sort_field: request.sort_order or "desc"}
        
        logger.info(f"Calling search_multiple_categories with {len(request.category_ids)} category IDs")
        
        # Perform multi-category search
        response = search_multiple_categories(
            category_ids=request.category_ids,
            shop_type=request.shop_type,
            page=request.page,
            page_size=request.page_size,
            max_products_per_category=request.max_products_per_category,
            price_start=request.price_start,
            price_end=request.price_end,
            sort=sort_param,
            region_opp=request.region_opp,
            filter=request.filter,
            request_timeout_seconds=30,
        )
        
        logger.info(f"search_multiple_categories returned: success={response.get('success')}, has_data={'data' in response}")
        
        if not response.get("success", False):
            error_msg = response.get("error", "Unknown error")
            logger.error(f"Multi-category search failed: {error_msg}")
            return ProductResponse(
                success=False,
                error=error_msg,
                message=f"Failed to search products across categories: {error_msg}"
            )
        
        # Extract products from response structure
        # Response structure: {success: bool, data: {result: {products: [], total: int, ...}}}
        products = []
        total = 0
        
        try:
            data = response.get("data", {})
            if isinstance(data, dict):
                result = data.get("result", {})
                if isinstance(result, dict):
                    products = result.get("products", [])
                    total = result.get("total", len(products))
                elif isinstance(result, list):
                    # Sometimes result might be directly a list
                    products = result
                    total = len(products)
            elif isinstance(data, list):
                # Sometimes data might be directly a list
                products = data
                total = len(products)
        except Exception as e:
            logger.error(f"Error extracting products from response: {e}")
            logger.error(f"Response structure: {response}")
            products = []
            total = 0
        
        logger.info(f"Extracted {len(products)} products from multi-category search response")
        
        # Apply quality filters if specified
        original_count = len(products)
        if products and (request.min_rating is not None or request.min_repurchase_rate is not None or request.min_monthly_sales is not None):
            filtered_products = []
            for product in products:
                # Filter by minimum rating
                if request.min_rating is not None:
                    rating = product.get("tradeScore") or product.get("rating") or product.get("trade_score") or 0
                    try:
                        rating_float = float(rating) if rating else 0
                        if rating_float < request.min_rating:
                            continue
                    except (ValueError, TypeError):
                        continue
                
                # Filter by minimum repurchase rate
                if request.min_repurchase_rate is not None:
                    repurchase = product.get("repurchaseRate") or product.get("rePurchaseRate") or product.get("repurchase_rate") or 0
                    try:
                        repurchase_float = float(repurchase) if repurchase else 0
                        if repurchase_float < request.min_repurchase_rate:
                            continue
                    except (ValueError, TypeError):
                        continue
                
                # Filter by minimum monthly sales
                if request.min_monthly_sales is not None:
                    sales = product.get("monthSold") or product.get("month_sold") or product.get("monthSoldCount") or 0
                    try:
                        sales_int = int(sales) if sales else 0
                        if sales_int < request.min_monthly_sales:
                            continue
                    except (ValueError, TypeError):
                        continue
                
                filtered_products.append(product)
            
            products = filtered_products
            logger.info(f"Applied quality filters: {len(products)}/{original_count} products remaining")
            if request.min_rating is not None:
                logger.info(f"  - Min rating: {request.min_rating}")
            if request.min_repurchase_rate is not None:
                logger.info(f"  - Min repurchase rate: {request.min_repurchase_rate}%")
            if request.min_monthly_sales is not None:
                logger.info(f"  - Min monthly sales: {request.min_monthly_sales}")
        
        # If no products found but search was successful, still return success
        if len(products) == 0 and response.get("success", False):
            logger.warning(f"No products found for categories: {request.category_ids} (after filtering: {original_count} -> {len(products)})")
            summary = response.get("data", {}).get("result", {}).get("summary", {})
            if summary:
                logger.info(f"Search summary: {summary}")
        
        # Enrich products with detail information (images, descriptions, etc.)
        # This is the same enrichment step that keyword search uses
        try:
            if products:
                # Enrich up to 50 products with full detail information
                enrich_limit = min(len(products), 50)
                logger.info(f"Enriching {enrich_limit} products with detail information...")
                enrich_products_with_detail(
                    products,
                    get_detail_fn=lambda **kwargs: get_product_detail(
                        goods_id=kwargs.get("goods_id"),
                        shop_type=kwargs.get("shop_type", request.shop_type),
                        request_timeout_seconds=30,
                        normalize=True,
                    ),
                    shop_type=request.shop_type,
                    request_timeout_seconds=30,
                    limit=enrich_limit,
                )
                logger.info(f"Successfully enriched {enrich_limit} products with detail information")
        except Exception as e:
            logger.warning(f"Enrich step failed (products will still be returned): {e}")
            import traceback
            logger.debug(f"Enrich error traceback: {traceback.format_exc()}")
        
        # Save to database if requested
        saved_count = 0
        if request.save_to_db and products:
            logger.info(f"Attempting to save {len(products)} products to database (save_to_db={request.save_to_db})")
            try:
                saved_count = save_products_to_db(products, keyword="multi-category-search")
                logger.info(f"Successfully saved {saved_count} products to database (out of {len(products)} products)")
                if saved_count == 0:
                    logger.warning(f"Warning: save_products_to_db returned 0 saved products, but {len(products)} products were provided")
                add_log("success", "Multi-category search products saved", f"Saved {saved_count} products from {len(request.category_ids)} categories", "database")
            except Exception as e:
                import traceback
                error_trace = traceback.format_exc()
                logger.error(f"Failed to save products to database: {e}")
                logger.error(f"Error traceback: {error_trace}")
                add_log("error", "Database save failed", f"Error: {str(e)}", "database")
                try:
                    logger.info("Attempting to fix schema and retry save...")
                    fix_products_origin_schema()
                    saved_count = save_products_to_db(products, keyword="multi-category-search")
                    logger.info(f"Fixed schema and saved {saved_count} products to database")
                    add_log("success", "Schema fixed and products saved", f"Saved {saved_count} products after schema fix", "database")
                except Exception as e2:
                    import traceback
                    error_trace2 = traceback.format_exc()
                    logger.error(f"Failed to save products even after schema fix: {e2}")
                    logger.error(f"Error traceback: {error_trace2}")
                    add_log("error", "Database save failed after schema fix", f"Error: {str(e2)}", "database")
        else:
            if not request.save_to_db:
                logger.info(f"Skipping database save because save_to_db={request.save_to_db}")
            if not products:
                logger.info(f"Skipping database save because products list is empty")
        
        # Log the search operation
        add_log(
            "success" if len(products) > 0 else "info",
            "Multi-category search completed",
            f"Searched {len(request.category_ids)} categories, found {len(products)} products, saved {saved_count} to database",
            "search"
        )
        
        return ProductResponse(
            success=True,
            data=products,
            total=total,
            message=f"Found {len(products)} products across {len(request.category_ids)} categories, saved {saved_count} to database"
        )
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Multi-category search failed: {e}")
        logger.error(f"Traceback: {error_trace}")
        add_log("error", "Multi-category search exception", f"Error: {str(e)}", "search")
        return ProductResponse(
            success=False,
            error=str(e),
            message=f"Multi-category search failed: {str(e)}",
            data=[],
            total=0
        )

@app.get("/api/test/price-filter")
async def test_price_filter():
    """
    Test price filtering functionality with sample data
    """
    try:
        logger.info("Testing price filter functionality...")
        
        # Test with a simple search and price range
        response = keyword_search_products(
            keywords="dress",
            shop_type="1688",
            page=1,
            page_size=10,
            price_start="10",
            price_end="100",
            request_timeout_seconds=15,
        )
        
        if response and response.get("success", False):
            parsed_response = parse_keyword_search_response(response)
            products = parsed_response.get("data", {}).get("result", {}).get("products", [])
            
            # Apply our post-processing price filter
            filtered_products = []
            for product in products:
                product_price = product.get("goodsPrice")
                if product_price is not None:
                    try:
                        price_val = float(product_price)
                        if 10 <= price_val <= 100:
                            filtered_products.append(product)
                    except (ValueError, TypeError):
                        pass
            
            return {
                "success": True,
                "message": f"Price filter test completed",
                "original_products": len(products),
                "filtered_products": len(filtered_products),
                "price_range": "10-100",
                "sample_prices": [float(p.get("goodsPrice", 0)) for p in products[:5] if p.get("goodsPrice")],
                "filtered_sample_prices": [float(p.get("goodsPrice", 0)) for p in filtered_products[:5] if p.get("goodsPrice")]
            }
        else:
            return {
                "success": False,
                "message": "Price filter test failed - API search failed",
                "error": response.get("error", "Unknown error") if response else "No response received"
            }
            
    except Exception as e:
        logger.error(f"Price filter test failed: {e}")
        return {
            "success": False,
            "message": "Price filter test failed",
            "error": str(e)
        }


@app.get("/api/test/rakumart-connection")
async def test_rakumart_connection():
    """
    Test the Rakumart API connection with a simple search
    """
    try:
        logger.info("Testing Rakumart API connection...")
        
        # Test with a simple keyword search
        response = keyword_search_products(
            keywords="test",
            shop_type="1688",
            page=1,
            page_size=5,
            request_timeout_seconds=10,
        )
        
        if response and response.get("success", False):
            parsed_response = parse_keyword_search_response(response)
            products = parsed_response.get("data", {}).get("result", {}).get("products", [])
            
            return {
                "success": True,
                "message": f"API connection successful. Found {len(products)} test products.",
                "api_response_structure": list(response.keys()) if isinstance(response, dict) else str(type(response)),
                "products_found": len(products)
            }
        else:
            return {
                "success": False,
                "message": "API connection failed",
                "error": response.get("error", "Unknown error") if response else "No response received"
            }
            
    except Exception as e:
        logger.error(f"Rakumart connection test failed: {e}")
        return {
            "success": False,
            "message": "Connection test failed",
            "error": str(e)
        }


@app.post("/api/database/cleanup-empty-records")
async def cleanup_empty_records_endpoint():
    """
    Clean up empty records from the database
    """
    try:
        from modules.db import cleanup_empty_records
        deleted_count = cleanup_empty_records()
        
        add_log("info", f"Database cleanup completed", f"Removed {deleted_count} empty records", "database")
        
        return {
            "success": True,
            "message": f"Cleaned up {deleted_count} empty records",
            "deleted_count": deleted_count
        }
        
    except Exception as e:
        logger.error(f"Failed to cleanup empty records: {e}")
        add_log("error", "Database cleanup failed", str(e), "database")
        return {
            "success": False,
            "error": str(e),
            "message": "Failed to cleanup empty records"
        }


@app.delete("/api/products/{product_id}")
async def delete_product(product_id: str):
    """
    Delete a product from the database by product_id
    """
    try:
        logger.info(f"Deleting product: {product_id}")
        
        # Import the db module
        from modules.db import get_db_connection
        
        # Get database connection
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            # Delete the product
            cur.execute("DELETE FROM products_origin WHERE product_id = %s", (product_id,))
            deleted_count = cur.rowcount
            
            if deleted_count == 0:
                logger.warning(f"No product found with id: {product_id}")
                add_log("warning", f"Product not found", f"Product ID: {product_id}", "delete")
                conn.close()
                return {
                    "success": False,
                    "error": "Product not found",
                    "message": f"No product found with id: {product_id}"
                }
            
            # Also delete from product_management
            try:
                pm_deleted = delete_product_management_by_item_numbers([product_id])
                logger.info(f"Also removed {pm_deleted} from product_management for item_number={product_id}")
            except Exception as e2:
                logger.warning(f"Failed to delete from product_management for {product_id}: {e2}")
            
            conn.commit()
            conn.close()
            
            logger.info(f"Successfully deleted product: {product_id}")
            add_log("success", "Product deleted", f"Product ID: {product_id}", "delete")
            
            return {
                "success": True,
                "message": f"Product {product_id} deleted successfully",
                "deleted_count": deleted_count
            }
            
        except Exception as e:
            logger.error(f"Database error while deleting product: {e}")
            conn.close()
            return {
                "success": False,
                "error": str(e),
                "message": "Failed to delete product"
            }
            
    except Exception as e:
        logger.error(f"Failed to delete product: {e}")
        add_log("error", "Product deletion failed", f"Product ID: {product_id}, Error: {str(e)}", "delete")
        return {
            "success": False,
            "error": str(e),
            "message": "Failed to delete product"
        }


@app.post("/api/products/batch-delete")
async def delete_products_batch(request: BatchDeleteRequest):
    """
    Delete multiple products from the database by their product_ids
    """
    try:
        product_ids = request.product_ids
        logger.info(f"Batch deleting {len(product_ids)} products")
        
        # Import the db module
        from modules.db import get_db_connection
        
        # Get database connection
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            # Delete all products at once
            cur.execute("DELETE FROM products_origin WHERE product_id = ANY(%s)", (product_ids,))
            deleted_count = cur.rowcount
            
            # Also delete from product_management
            try:
                pm_deleted = delete_product_management_by_item_numbers(product_ids)
                logger.info(f"Also removed {pm_deleted} from product_management")
            except Exception as e2:
                logger.warning(f"Failed to delete from product_management: {e2}")
            
            conn.commit()
            conn.close()
            
            logger.info(f"Successfully deleted {deleted_count} products")
            add_log("success", "Batch product deletion", f"Deleted {deleted_count} products", "delete")
            
            return {
                "success": True,
                "message": f"Deleted {deleted_count} products successfully",
                "deleted_count": deleted_count
            }
            
        except Exception as e:
            logger.error(f"Database error while batch deleting products: {e}")
            conn.close()
            return {
                "success": False,
                "error": str(e),
                "message": "Failed to delete products"
            }
            
    except Exception as e:
        logger.error(f"Failed to batch delete products: {e}")
        add_log("error", "Batch product deletion failed", f"Error: {str(e)}", "delete")
        return {
            "success": False,
            "error": str(e),
            "message": "Failed to delete products"
        }


@app.delete("/api/product-management/{item_number}")
async def delete_product_management_item(item_number: str):
    """Delete a product from the product_management table by item_number."""
    try:
        deleted = delete_product_management_by_item_numbers([item_number])
        if deleted > 0:
            add_log("success", "Product management item deleted", f"Item number: {item_number}", "product_management")
            return {
                "success": True,
                "message": f"Deleted product_management item {item_number}",
                "deleted_count": deleted,
            }
        else:
            return {
                "success": False,
                "error": "Item not found",
                "message": f"Product_management item {item_number} was not found",
            }
    except Exception as e:
        logger.error(f"Failed to delete product_management item {item_number}: {e}")
        add_log("error", "Product management item deletion failed", str(e), "product_management")
        return {
            "success": False,
            "error": str(e),
            "message": "Failed to delete product_management item",
        }


@app.post("/api/product-management/batch-delete")
async def delete_product_management_items(request: ProductManagementBatchDeleteRequest):
    """Delete multiple products from the product_management table."""
    try:
        item_numbers = [item for item in request.item_numbers if item]
        if not item_numbers:
            return {
                "success": False,
                "error": "No item numbers provided",
                "message": "Please provide at least one item_number",
            }

        deleted = delete_product_management_by_item_numbers(item_numbers)
        add_log(
            "success",
            "Batch delete product management items",
            f"Deleted {deleted} items",
            "product_management",
        )
        return {
            "success": True,
            "message": f"Deleted {deleted} product_management items",
            "deleted_count": deleted,
        }
    except Exception as e:
        logger.error(f"Failed to batch delete product_management items: {e}")
        add_log("error", "Batch delete product management items failed", str(e), "product_management")
        return {
            "success": False,
            "error": str(e),
            "message": "Failed to delete product_management items",
        }


@app.get("/api/products", response_model=ProductResponse)
async def get_products_from_database(
    limit: int = 50,
    offset: int = 0,
    keyword: Optional[str] = None
):
    """
    Get products from the database
    """
    try:
        logger.info(f"Getting products from database: limit={limit}, offset={offset}, keyword={keyword}")
        
        products = get_products_from_db(limit=limit, offset=offset, keyword=keyword)
        
        return ProductResponse(
            success=True,
            data=products,
            total=len(products),
            message=f"Retrieved {len(products)} products from database"
        )
    except Exception as e:
        logger.error(f"Failed to get products from database: {e}")
        return ProductResponse(
            success=False,
            error=str(e),
            message="Failed to retrieve products from database"
        )

@app.patch("/api/product-management/{item_number}/hide-item")
async def update_product_hide_item_endpoint(item_number: str, hide_item: bool):
    """
    Update the hide_item field for a single product.
    
    Args:
        item_number: The product item_number
        hide_item: Boolean value (True = hidden, False = visible)
    """
    try:
        logger.info(f"Updating hide_item for product {item_number} to {hide_item}")
        success = update_product_hide_item(item_number, hide_item)
        
        if success:
            add_log("info", f"Product visibility updated", f"Product {item_number}: {'hidden' if hide_item else 'visible'}", "product_management")
            return {
                "success": True,
                "message": f"Product {item_number} visibility updated to {'hidden' if hide_item else 'visible'}"
            }
        else:
            return {
                "success": False,
                "error": "Product not found",
                "message": f"No product found with item_number: {item_number}"
            }
    except Exception as e:
        logger.error(f"Failed to update hide_item for product {item_number}: {e}")
        add_log("error", "Failed to update product visibility", str(e), "product_management")
        return {
            "success": False,
            "error": str(e),
            "message": f"Failed to update product visibility: {str(e)}"
        }


@app.patch("/api/product-management/batch-hide-item")
async def update_products_hide_item_batch_endpoint(request: UpdateHideItemRequest):
    """
    Update the hide_item field for multiple products.
    """
    try:
        if not request.item_numbers:
            return {
                "success": False,
                "error": "No item_numbers provided",
                "message": "At least one item_number is required"
            }
        
        logger.info(f"Updating hide_item for {len(request.item_numbers)} products to {request.hide_item}")
        updated_count = update_products_hide_item_batch(request.item_numbers, request.hide_item)
        
        add_log("info", f"Batch product visibility updated", f"{updated_count} products set to {'hidden' if request.hide_item else 'visible'}", "product_management")
        return {
            "success": True,
            "message": f"Updated {updated_count} products to {'hidden' if request.hide_item else 'visible'}",
            "updated_count": updated_count
        }
    except Exception as e:
        logger.error(f"Failed to update hide_item for products: {e}")
        add_log("error", "Failed to update batch product visibility", str(e), "product_management")
        return {
            "success": False,
            "error": str(e),
            "message": f"Failed to update product visibility: {str(e)}"
        }


@app.patch("/api/product-management/all-hide-item")
async def update_all_products_hide_item_endpoint(hide_item: bool):
    """
    Update the hide_item field for ALL products in the database.
    
    Args:
        hide_item: Boolean value (True = hidden, False = visible)
    """
    try:
        logger.info(f"Updating hide_item for ALL products to {hide_item}")
        updated_count = update_all_products_hide_item(hide_item)
        
        add_log("info", f"All products visibility updated", f"All {updated_count} products set to {'hidden' if hide_item else 'visible'}", "product_management")
        return {
            "success": True,
            "message": f"Updated ALL {updated_count} products to {'hidden' if hide_item else 'visible'}",
            "updated_count": updated_count
        }
    except Exception as e:
        logger.error(f"Failed to update hide_item for all products: {e}")
        add_log("error", "Failed to update all products visibility", str(e), "product_management")
        return {
            "success": False,
            "error": str(e),
            "message": f"Failed to update all products visibility: {str(e)}"
        }


@app.patch("/api/product-management/{item_number}/settings")
async def update_product_settings_endpoint(item_number: str, request: dict):
    """
    Update product settings in the product_management table.
    
    Request body can include:
    - item_type: string (NORMAL, PRE_ORDER, BUYING_CLUB)
    - genre_id: string or null (100000-999999)
    - tags: array of numbers (5000000-9999999)
    - unlimited_inventory_flag: boolean
    - features: object with searchVisibility, inventoryDisplay, review
    - payment: object with taxIncluded, taxRate, cashOnDeliveryFeeIncluded
    """
    try:
        logger.info(f"Updating settings for product {item_number}")
        
        success = update_product_management_settings(
            item_number=item_number,
            item_type=request.get("item_type"),
            genre_id=request.get("genre_id"),
            tags=request.get("tags"),
            unlimited_inventory_flag=request.get("unlimited_inventory_flag"),
            features=request.get("features"),
            payment=request.get("payment"),
            title=request.get("title"),
            normal_delivery_date_id=request.get("normalDeliveryDateId"),
        )
        
        if success:
            add_log("info", f"Product settings updated", f"Product {item_number} settings updated successfully", "product_management")
            return {
                "success": True,
                "message": f"Product {item_number} settings updated successfully"
            }
        else:
            return {
                "success": False,
                "error": "Product not found",
                "message": f"No product found with item_number: {item_number}"
            }
    except Exception as e:
        logger.error(f"Failed to update product settings for {item_number}: {e}")
        add_log("error", "Failed to update product settings", str(e), "product_management")
        return {
            "success": False,
            "error": str(e),
            "message": f"Failed to update product settings: {str(e)}"
        }


@app.patch("/api/product-management/{item_number}/sku")
async def update_product_sku_endpoint(item_number: str, request: Optional[dict] = Body(None)):
    """
    Update SKU data (variant_selectors and/or variants) for a specific product.
    
    Request body can include:
    - variant_selectors: array of selector definitions
      [{ key: string, displayName: string, values: [{ displayValue: string }] }]
    - variants: object of variants keyed by skuId
      { skuId: { selectorValues, standardPrice, shipping, features, ... } }
    """
    try:
        logger.info(f"📦 [SKU API] Received SKU update request for product {item_number}")
        logger.info(f"📦 [SKU API] Request body type: {type(request)}")
        
        if request is None or not isinstance(request, dict):
            logger.warning(f"Empty request body for product {item_number}")
            return {
                "success": False,
                "error": "Invalid request",
                "message": "Request body is required"
            }
        
        variant_selectors = request.get("variant_selectors")
        variants = request.get("variants")
        
        # Log received data
        if variant_selectors is not None:
            logger.info(f"📋 [SKU API] Received variant_selectors: {len(variant_selectors)} selector(s)")
            for idx, selector in enumerate(variant_selectors):
                selector_key = selector.get("key", "unknown")
                selector_display = selector.get("displayName", selector_key)
                values_count = len(selector.get("values", []))
                logger.info(f"   Selector {idx + 1}: key='{selector_key}', displayName='{selector_display}', values={values_count}")
                for v_idx, value in enumerate(selector.get("values", [])):
                    logger.info(f"      Value {v_idx + 1}: '{value.get('displayValue', '')}'")
        else:
            logger.info("📋 [SKU API] No variant_selectors in request")
        
        if variants is not None:
            logger.info(f"📦 [SKU API] Received variants: {len(variants)} variant(s)")
            for sku_id, variant_data in variants.items():
                logger.info(f"   Variant SKU: {sku_id}")
                if variant_data.get("selectorValues"):
                    logger.info(f"      Selector Values: {variant_data['selectorValues']}")
                if variant_data.get("standardPrice"):
                    logger.info(f"      Price: ¥{variant_data['standardPrice']}")
                if variant_data.get("articleNumber"):
                    logger.info(f"      Article Number: {variant_data['articleNumber']}")
                if variant_data.get("shipping"):
                    logger.info(f"      Shipping: {variant_data['shipping']}")
                if variant_data.get("features"):
                    logger.info(f"      Features: {variant_data['features']}")
        else:
            logger.info("📦 [SKU API] No variants in request")
        
        # Allow None values (means don't update that field), but log if both are None
        if variant_selectors is None and variants is None:
            logger.warning(f"No SKU data provided for product {item_number}")
            return {
                "success": False,
                "error": "No SKU data provided",
                "message": "Please provide variant_selectors or variants to update"
            }
        
        # Validate data types
        if variant_selectors is not None and not isinstance(variant_selectors, list):
            return {
                "success": False,
                "error": "Invalid variant_selectors",
                "message": "variant_selectors must be an array"
            }
        
        if variants is not None and not isinstance(variants, dict):
            return {
                "success": False,
                "error": "Invalid variants",
                "message": "variants must be an object"
            }
        
        success = update_product_sku_data(
            item_number=item_number,
            variant_selectors=variant_selectors,
            variants=variants,
        )
        
        if success:
            add_log("info", "Product SKU data updated", f"Product {item_number} SKU data updated successfully", "product_management")
            return {
                "success": True,
                "message": f"Product {item_number} SKU data updated successfully"
            }
        else:
            return {
                "success": False,
                "error": "Product not found",
                "message": f"No product found with item_number: {item_number}"
            }
    except ValueError as e:
        logger.error(f"Validation error updating SKU data for {item_number}: {e}")
        add_log("error", "Failed to update SKU data - validation error", str(e), "product_management")
        return {
            "success": False,
            "error": "Validation error",
            "message": f"Invalid data format: {str(e)}"
        }
    except Exception as e:
        logger.error(f"Failed to update SKU data for {item_number}: {e}", exc_info=True)
        add_log("error", "Failed to update SKU data", str(e), "product_management")
        return {
            "success": False,
            "error": str(e),
            "message": f"Failed to update SKU data: {str(e)}"
        }


@app.patch("/api/product-management/{item_number}/variant/{sku_id}")
async def update_single_variant_endpoint(item_number: str, sku_id: str, request: dict):
    """
    Update a single variant within a product's variants field.
    
    Request body contains the variant fields to update:
    - standardPrice: string (price in yen)
    - selectorValues: object { selectorKey: value }
    - shipping: object { postageIncluded: boolean, ... }
    - features: object { restockNotification: boolean, ... }
    """
    try:
        logger.info(f"Updating variant {sku_id} for product {item_number}")
        
        if not request:
            return {
                "success": False,
                "error": "No variant data provided",
                "message": "Please provide variant data to update"
            }
        
        success = update_single_variant(
            item_number=item_number,
            sku_id=sku_id,
            variant_data=request,
        )
        
        if success:
            add_log("info", "Variant updated", f"Variant {sku_id} for product {item_number} updated successfully", "product_management")
            return {
                "success": True,
                "message": f"Variant {sku_id} updated successfully"
            }
        else:
            return {
                "success": False,
                "error": "Variant or product not found",
                "message": f"Could not find variant {sku_id} in product {item_number}"
            }
    except Exception as e:
        logger.error(f"Failed to update variant {sku_id} for {item_number}: {e}")
        add_log("error", "Failed to update variant", str(e), "product_management")
        return {
            "success": False,
            "error": str(e),
            "message": f"Failed to update variant: {str(e)}"
        }


@app.post("/api/product-management/register", response_model=DatabaseResponse)
async def register_products_to_management(request: RegisterProductsRequest):
    """Register selected products (by product_id) into product_management table."""
    try:
        if not request.product_ids:
            return DatabaseResponse(success=False, error="No product_ids provided", message="No products to register")
        
        # Filter product_ids based on rakuten_registration_status
        from modules.db import get_product_management_by_item_number
        filtered_product_ids = []
        skipped_count = 0
        
        for product_id in request.product_ids:
            # Check if product exists in product_management and get its status
            product_data = get_product_management_by_item_number(product_id)
            if product_data:
                current_status = product_data.get("rakuten_registration_status")
                # Only process if status is NULL, empty, 'onsale', 'true', or 'false'
                if current_status not in [None, "", "onsale", "true", "false"]:
                    logger.info(f"Skipping product {product_id}: rakuten_registration_status is '{current_status}' (only NULL, '', 'onsale', 'true', or 'false' are allowed)")
                    skipped_count += 1
                    continue
            # If product doesn't exist in product_management, allow registration (status is NULL)
            filtered_product_ids.append(product_id)
        
        if not filtered_product_ids:
            return DatabaseResponse(
                success=False, 
                error="No products to register after filtering", 
                message=f"すべての商品がスキップされました（rakuten_registration_statusが条件を満たしていません）"
            )
        
        if skipped_count > 0:
            logger.info(f"Filtered {skipped_count} products based on rakuten_registration_status")
        
        saved = upsert_product_management_from_origin_ids(filtered_product_ids)
        add_log("success", "Products registered to product_management", f"Saved {saved} items (skipped {skipped_count})", "database")
        return DatabaseResponse(
            success=True, 
            saved_count=saved, 
            message=f"Registered {saved} products" + (f" ({skipped_count} skipped)" if skipped_count > 0 else "")
        )
    except Exception as e:
        logger.error(f"Failed to register products: {e}")
        add_log("error", "Register to product_management failed", str(e), "database")
        return DatabaseResponse(success=False, error=str(e), message="Failed to register products")


@app.post("/api/product-management/update-variants", response_model=DatabaseResponse)
async def update_variants_only(request: RegisterProductsRequest):
    """
    Update ONLY variant_selectors and variants for products already registered in product_management.
    Source data is products_origin.detail_json; no image processing or other fields.
    Only updates products where rakuten_registration_status is NULL, empty, 'onsale', 'true', or 'false'.
    """
    try:
        if not request.product_ids:
            return DatabaseResponse(success=False, error="No product_ids provided", message="No products to update")
        
        # Filter product_ids based on rakuten_registration_status
        from modules.db import get_product_management_by_item_number
        filtered_product_ids = []
        skipped_count = 0
        
        for product_id in request.product_ids:
            # Check if product exists in product_management and get its status
            product_data = get_product_management_by_item_number(product_id)
            if not product_data:
                logger.warning(f"Product {product_id} not found in product_management, skipping")
                skipped_count += 1
                continue
            
            current_status = product_data.get("rakuten_registration_status")
            # Only process if status is NULL, empty, 'onsale', 'true', or 'false'
            if current_status not in [None, "", "onsale", "true", "false"]:
                logger.info(f"Skipping product {product_id}: rakuten_registration_status is '{current_status}' (only NULL, '', 'onsale', 'true', or 'false' are allowed)")
                skipped_count += 1
                continue
            
            filtered_product_ids.append(product_id)
        
        if not filtered_product_ids:
            return DatabaseResponse(
                success=False, 
                error="No products to update after filtering", 
                message=f"すべての商品がスキップされました（rakuten_registration_statusが条件を満たしていません）"
            )
        
        if skipped_count > 0:
            logger.info(f"Filtered {skipped_count} products based on rakuten_registration_status")
        
        updated = update_variant_selectors_and_variants(filtered_product_ids)
        add_log("success", "Variants updated in product_management", f"Updated {updated} items (skipped {skipped_count})", "database")
        return DatabaseResponse(
            success=True, 
            saved_count=updated, 
            message=f"Updated variants for {updated} products" + (f" ({skipped_count} skipped)" if skipped_count > 0 else "")
        )
    except Exception as e:
        logger.error(f"Failed to update variants: {e}")
        add_log("error", "Update variants failed", str(e), "database")
        return DatabaseResponse(success=False, error=str(e), message="Failed to update variants")


@app.post("/api/product-management/register-multiple-to-rakuten")
async def register_multiple_products_to_rakuten(request: RegisterMultipleToRakutenRequest):
    """Register multiple products from product_management table to Rakuten API sequentially."""
    try:
        item_numbers = request.item_numbers
        if not item_numbers or len(item_numbers) == 0:
            return {
                "success": False,
                "error": "item_numbers is required",
                "message": "Please provide at least one item_number"
            }
        
        from modules.db import update_rakuten_registration_status, get_product_management_by_item_number
        
        results = []
        success_count = 0
        failure_count = 0
        skipped_count = 0
        
        # Process each product sequentially
        for idx, item_number in enumerate(item_numbers, 1):
            if not item_number:
                continue
            
            # Check rakuten_registration_status before processing
            product_data = get_product_management_by_item_number(item_number)
            if not product_data:
                logger.warning(f"Product {item_number} not found in database, skipping")
                results.append({
                    "item_number": item_number,
                    "success": False,
                    "skipped": True,
                    "error": "Product not found in database",
                    "message": "Product not found in database"
                })
                skipped_count += 1
                continue
            
            current_status = product_data.get("rakuten_registration_status")
            # Only process if status is NULL, empty, 'onsale', 'true', or 'false'
            if current_status not in [None, "", "onsale", "true", "false"]:
                logger.info(f"Skipping product {item_number}: rakuten_registration_status is '{current_status}' (only NULL, '', 'onsale', 'true', or 'false' are allowed)")
                results.append({
                    "item_number": item_number,
                    "success": False,
                    "skipped": True,
                    "error": f"Product registration status '{current_status}' does not allow registration",
                    "message": f"登録ステータスが '{current_status}' のため、登録をスキップしました"
                })
                skipped_count += 1
                continue
                
            try:
                logger.info(f"Registering product {idx}/{len(item_numbers)}: {item_number}")
                
                # Register product to Rakuten using the rakuten_product module
                result = register_product_from_product_management(item_number)
                
                if result.get("success"):
                    logger.info(f"Successfully registered product {item_number} to Rakuten")
                    add_log("success", f"Product {item_number} registered to Rakuten", result.get("message", ""), "rakuten")
                    
                    # Update registration status in database (success)
                    update_rakuten_registration_status(item_number, "true")
                    
                    results.append({
                        "item_number": item_number,
                        "success": True,
                        "message": result.get("message", "Product registered successfully")
                    })
                    success_count += 1
                else:
                    error_msg = result.get("error", "Unknown error")
                    error_details = format_error_message(result)
                    
                    # Log detailed error information
                    logger.error(f"Failed to register product {item_number} to Rakuten: {error_msg}")
                    if result.get("error_data"):
                        logger.error(f"Rakuten API Error Data: {json.dumps(result.get('error_data'), indent=2, ensure_ascii=False)}")
                    if result.get("error_text"):
                        logger.error(f"Rakuten API Error Text: {result.get('error_text')}")
                    if result.get("status_code"):
                        logger.error(f"Rakuten API Status Code: {result.get('status_code')}")
                    
                    add_log("error", f"Failed to register product {item_number} to Rakuten", error_msg, "rakuten")
                    
                    # Update registration status in database (failed)
                    update_rakuten_registration_status(item_number, "false")
                    
                    results.append({
                        "item_number": item_number,
                        "success": False,
                        "error": error_msg,
                        "error_details": error_details,
                        "message": f"Failed to register product to Rakuten: {error_msg}"
                    })
                    failure_count += 1
                    
            except Exception as e:
                error_msg = str(e)
                logger.error(f"Exception while registering product {item_number} to Rakuten: {error_msg}")
                add_log("error", f"Exception while registering product {item_number} to Rakuten", error_msg, "rakuten")
                
                # Update registration status in database (failed)
                try:
                    update_rakuten_registration_status(item_number, "false")
                except Exception as db_error:
                    logger.error(f"Failed to update registration status: {db_error}")
                
                results.append({
                    "item_number": item_number,
                    "success": False,
                    "error": error_msg,
                    "message": f"Failed to register product to Rakuten: {error_msg}"
                })
                failure_count += 1
        
        # Return aggregated results
        total_count = len(item_numbers)
        overall_success = failure_count == 0
        
        return {
            "success": overall_success,
            "total_count": total_count,
            "success_count": success_count,
            "failure_count": failure_count,
            "skipped_count": skipped_count,
            "results": results,
            "message": f"Registered {success_count}/{total_count} products successfully" + (f" ({failure_count} failed)" if failure_count > 0 else "") + (f" ({skipped_count} skipped)" if skipped_count > 0 else "")
        }
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Exception while registering multiple products to Rakuten: {error_msg}")
        add_log("error", "Exception while registering multiple products to Rakuten", error_msg, "rakuten")
        
        return {
            "success": False,
            "error": error_msg,
            "message": f"Failed to register products to Rakuten: {error_msg}"
        }


@app.post("/api/product-management/{item_number}/upload-images-to-rakuten")
async def upload_product_images_to_rakuten(item_number: str):
    """
    Upload product images to Rakuten Cabinet.
    
    Uses the images field from product_management table to construct URLs
    and upload them to Rakuten using batch upload.
    """
    try:
        from modules.db import get_product_management_by_item_number
        from modules.upload_file import batch_upload_images
        
        # Get product data
        product_data = get_product_management_by_item_number(item_number)
        if not product_data:
            return {
                "success": False,
                "error": f"Product with item_number '{item_number}' not found",
                "message": f"No product found with item_number: {item_number}"
            }
        
        # Get images field
        images = product_data.get("images")
        if not images:
            return {
                "success": False,
                "error": f"No images found for product '{item_number}'",
                "message": f"Product {item_number} has no images to upload"
            }
        
        # Parse images if it's a string
        if isinstance(images, str):
            try:
                images = json.loads(images)
            except json.JSONDecodeError:
                return {
                    "success": False,
                    "error": f"Invalid images JSON format for product '{item_number}'",
                    "message": f"Invalid images data format"
                }
        
        if not isinstance(images, list):
            return {
                "success": False,
                "error": f"Images must be a list for product '{item_number}'",
                "message": f"Invalid images data structure"
            }
        
        if len(images) == 0:
            return {
                "success": False,
                "error": f"No images in images array for product '{item_number}'",
                "message": f"Product {item_number} has no images to upload"
            }
        
        # Construct URLs from images and extract image_key
        base_url = "https://licel-product-image.s3.ap-southeast-2.amazonaws.com/products"
        urls = []
        image_key = None
        
        for image in images:
            if not isinstance(image, dict):
                continue
            
            location = image.get('location')
            if not location:
                continue
            
            # Extract image_key from location (e.g., "/img01306503/01306503_1.jpg" -> "01306503")
            # Location format: /{image_key}/{image_key}_{number}.jpg or /img{image_key}/{image_key}_{number}.jpg
            if not image_key:
                # Ensure location starts with /
                clean_location = location if location.startswith("/") else f"/{location}"
                # Remove leading / and split by /
                parts = clean_location.lstrip("/").split("/")
                if parts and len(parts) > 0:
                    # First part is the directory name (e.g., "img01306503" or "01306503")
                    directory_part = parts[0]
                    # Remove 'img' prefix if present to get the actual image_key
                    if directory_part.startswith('img') and len(directory_part) > 3:
                        # Check if after 'img' it's all numeric (image_key)
                        after_img = directory_part[3:]
                        if after_img.isdigit():
                            # Use the actual image_key without prefix
                            image_key = after_img
                        else:
                            # Not a numeric image_key, use as-is
                            image_key = directory_part
                    else:
                        # No 'img' prefix, use as-is
                        image_key = directory_part
            
            # Ensure location starts with /
            clean_location = location if location.startswith("/") else f"/{location}"
            # Construct full URL
            full_url = f"{base_url}{clean_location}"
            urls.append(full_url)
        
        if len(urls) == 0:
            return {
                "success": False,
                "error": f"No valid image locations found for product '{item_number}'",
                "message": f"No valid URLs could be constructed from images"
            }
        
        logger.info(f"Uploading {len(urls)} images to Rakuten for product {item_number}")
        
        # Generate folder name from product title or item_number
        product_title = product_data.get("title") or item_number
        # Clean folder name (remove special characters, limit length)
        # Rakuten folder names should not contain special characters
        import re
        # Remove all special characters except alphanumeric, space, hyphen, underscore
        folder_name = re.sub(r'[^a-zA-Z0-9\s\-_]', '', str(product_title))
        # Remove leading/trailing spaces and replace multiple spaces with single space
        folder_name = re.sub(r'\s+', ' ', folder_name).strip()
        # Limit to 50 characters
        folder_name = folder_name[:50]
        # If empty after cleaning, use item_number
        if not folder_name:
            folder_name = f"Product_{item_number}"
        # Final validation: ensure it's not empty and doesn't exceed 50 chars
        if len(folder_name.encode('utf-8')) > 50:
            # Truncate by character count until it fits in 50 bytes
            while len(folder_name.encode('utf-8')) > 50 and len(folder_name) > 0:
                folder_name = folder_name[:-1]
        if not folder_name:
            folder_name = f"Product_{item_number}"
        
        # Use item_number as name prefix for images
        name_prefix = item_number
        
        # Use image_key as directory_name (extracted from image location)
        # If image_key not found, fallback to item_number
        if not image_key:
            image_key = item_number
        
        # Use image_key directly as directory_name (don't modify it - it should be like "01469590" or "677868580085")
        # The upload_file module will clean and validate it
        directory_name = image_key
        
        logger.info(f"Using directory_name (image_key): '{directory_name}' (extracted from image location)")
        
        # Upload images
        result = batch_upload_images(
            urls=urls,
            folder_name=folder_name,
            folder_id=None,  # Let it create the folder
            name_prefix=name_prefix,
            directory_name=directory_name,
            image_key=image_key  # Pass image_key for file_path_name generation
        )
        
        if result.get("success") or result.get("successful", 0) > 0:
            uploaded_count = result.get("successful", 0)
            failed_count = result.get("failed", 0)
            
            logger.info(f"Successfully uploaded {uploaded_count}/{len(urls)} images for product {item_number}")
            add_log("success", f"Uploaded images to Rakuten for product {item_number}", 
                   f"Uploaded {uploaded_count}/{len(urls)} images", "rakuten")
            
            # Update image_registration_status to true on successful upload
            from modules.db import update_product_registration_status
            try:
                update_product_registration_status(
                    item_number,
                    image_registration_status=True
                )
                logger.info(f"Updated image_registration_status to true for product {item_number}")
            except Exception as e:
                logger.warning(f"Failed to update image_registration_status for product {item_number}: {e}")
            
            return {
                "success": True,
                "message": f"Successfully uploaded {uploaded_count}/{len(urls)} images to Rakuten",
                "total": len(urls),
                "uploaded_count": uploaded_count,
                "failed_count": failed_count,
                "uploaded_files": result.get("uploaded_files", []),
                "folder_id": result.get("folder_id"),
                "folder_name": result.get("folder_name"),
                "errors": result.get("errors", [])
            }
        else:
            error_msg = result.get("error", "Upload failed")
            logger.error(f"Failed to upload images for product {item_number}: {error_msg}")
            add_log("error", f"Failed to upload images to Rakuten for product {item_number}", 
                   error_msg, "rakuten")
            
            return {
                "success": False,
                "error": error_msg,
                "message": f"Failed to upload images: {error_msg}",
                "errors": result.get("errors", []),
                "uploaded_files": result.get("uploaded_files", [])
            }
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Exception while uploading images for product {item_number}: {error_msg}")
        add_log("error", f"Exception while uploading images to Rakuten for product {item_number}", 
               error_msg, "rakuten")
        return {
            "success": False,
            "error": error_msg,
            "message": f"An unexpected error occurred: {error_msg}"
        }


@app.post("/api/product-management/upload-multiple-images-to-rakuten")
async def upload_multiple_images_to_rakuten(request: RegisterMultipleImagesToRakutenRequest):
    """Upload images for multiple products to Rakuten Cabinet sequentially."""
    try:
        item_numbers = request.item_numbers
        if not item_numbers or len(item_numbers) == 0:
            return {
                "success": False,
                "error": "item_numbers is required",
                "message": "Please provide at least one item_number"
            }
        
        from modules.db import get_product_management_by_item_number, update_product_registration_status
        from modules.upload_file import batch_upload_images
        import re
        
        results = []
        total_uploaded_images = 0
        total_failed_images = 0
        total_images = 0
        
        # Process each product sequentially
        for idx, item_number in enumerate(item_numbers, 1):
            if not item_number:
                continue
                
            try:
                logger.info(f"Uploading images for product {idx}/{len(item_numbers)}: {item_number}")
                
                # Get product data
                product_data = get_product_management_by_item_number(item_number)
                if not product_data:
                    results.append({
                        "item_number": item_number,
                        "success": False,
                        "error": f"Product with item_number '{item_number}' not found",
                        "message": f"No product found with item_number: {item_number}",
                        "uploaded_count": 0,
                        "failed_count": 0,
                        "total_count": 0
                    })
                    continue
                
                # Get images field
                images = product_data.get("images")
                if not images:
                    results.append({
                        "item_number": item_number,
                        "success": False,
                        "error": f"No images found for product '{item_number}'",
                        "message": f"Product {item_number} has no images to upload",
                        "uploaded_count": 0,
                        "failed_count": 0,
                        "total_count": 0
                    })
                    continue
                
                # Parse images if it's a string
                if isinstance(images, str):
                    try:
                        images = json.loads(images)
                    except json.JSONDecodeError:
                        results.append({
                            "item_number": item_number,
                            "success": False,
                            "error": f"Invalid images JSON format for product '{item_number}'",
                            "message": f"Invalid images data format",
                            "uploaded_count": 0,
                            "failed_count": 0,
                            "total_count": 0
                        })
                        continue
                
                if not isinstance(images, list) or len(images) == 0:
                    results.append({
                        "item_number": item_number,
                        "success": False,
                        "error": f"No images in images array for product '{item_number}'",
                        "message": f"Product {item_number} has no images to upload",
                        "uploaded_count": 0,
                        "failed_count": 0,
                        "total_count": 0
                    })
                    continue
                
                # Construct URLs from images and extract image_key
                base_url = "https://licel-product-image.s3.ap-southeast-2.amazonaws.com/products"
                urls = []
                image_key = None
                
                for image in images:
                    if not isinstance(image, dict):
                        continue
                    
                    location = image.get('location')
                    if not location:
                        continue
                    
                    # Extract image_key from location
                    if not image_key:
                        clean_location = location if location.startswith("/") else f"/{location}"
                        parts = clean_location.lstrip("/").split("/")
                        if parts and len(parts) > 0:
                            directory_part = parts[0]
                            if directory_part.startswith('img') and len(directory_part) > 3:
                                after_img = directory_part[3:]
                                if after_img.isdigit():
                                    image_key = after_img
                                else:
                                    image_key = directory_part
                            else:
                                image_key = directory_part
                    
                    clean_location = location if location.startswith("/") else f"/{location}"
                    full_url = f"{base_url}{clean_location}"
                    urls.append(full_url)
                
                if len(urls) == 0:
                    results.append({
                        "item_number": item_number,
                        "success": False,
                        "error": f"No valid image locations found for product '{item_number}'",
                        "message": f"No valid URLs could be constructed from images",
                        "uploaded_count": 0,
                        "failed_count": 0,
                        "total_count": 0
                    })
                    continue
                
                total_images += len(urls)
                
                # Generate folder name from product title or item_number
                product_title = product_data.get("title") or item_number
                folder_name = re.sub(r'[^a-zA-Z0-9\s\-_]', '', str(product_title))
                folder_name = re.sub(r'\s+', ' ', folder_name).strip()
                folder_name = folder_name[:50]
                if not folder_name:
                    folder_name = f"Product_{item_number}"
                if len(folder_name.encode('utf-8')) > 50:
                    while len(folder_name.encode('utf-8')) > 50 and len(folder_name) > 0:
                        folder_name = folder_name[:-1]
                if not folder_name:
                    folder_name = f"Product_{item_number}"
                
                name_prefix = item_number
                if not image_key:
                    image_key = item_number
                directory_name = image_key
                
                # Upload images
                result = batch_upload_images(
                    urls=urls,
                    folder_name=folder_name,
                    folder_id=None,
                    name_prefix=name_prefix,
                    directory_name=directory_name,
                    image_key=image_key
                )
                
                if result.get("success") or result.get("successful", 0) > 0:
                    uploaded_count = result.get("successful", 0)
                    failed_count = result.get("failed", 0)
                    
                    total_uploaded_images += uploaded_count
                    total_failed_images += failed_count
                    
                    logger.info(f"Successfully uploaded {uploaded_count}/{len(urls)} images for product {item_number}")
                    add_log("success", f"Uploaded images to Rakuten for product {item_number}", 
                           f"Uploaded {uploaded_count}/{len(urls)} images", "rakuten")
                    
                    # Update image_registration_status to true on successful upload
                    try:
                        update_product_registration_status(
                            item_number,
                            image_registration_status=True
                        )
                        logger.info(f"Updated image_registration_status to true for product {item_number}")
                    except Exception as e:
                        logger.warning(f"Failed to update image_registration_status for product {item_number}: {e}")
                    
                    results.append({
                        "item_number": item_number,
                        "success": True,
                        "message": f"Successfully uploaded {uploaded_count}/{len(urls)} images to Rakuten",
                        "uploaded_count": uploaded_count,
                        "failed_count": failed_count,
                        "total_count": len(urls),
                        "uploaded_files": result.get("uploaded_files", []),
                        "folder_id": result.get("folder_id"),
                        "folder_name": result.get("folder_name"),
                        "errors": result.get("errors", [])
                    })
                else:
                    error_msg = result.get("error", "Upload failed")
                    total_failed_images += len(urls)
                    
                    logger.error(f"Failed to upload images for product {item_number}: {error_msg}")
                    add_log("error", f"Failed to upload images to Rakuten for product {item_number}", 
                           error_msg, "rakuten")
                    
                    results.append({
                        "item_number": item_number,
                        "success": False,
                        "error": error_msg,
                        "message": f"Failed to upload images to Rakuten: {error_msg}",
                        "uploaded_count": 0,
                        "failed_count": len(urls),
                        "total_count": len(urls),
                        "errors": result.get("errors", [])
                    })
                    
            except Exception as e:
                error_msg = str(e)
                logger.error(f"Exception while uploading images for product {item_number}: {error_msg}")
                add_log("error", f"Exception while uploading images for product {item_number}", error_msg, "rakuten")
                
                results.append({
                    "item_number": item_number,
                    "success": False,
                    "error": error_msg,
                    "message": f"Failed to upload images to Rakuten: {error_msg}",
                    "uploaded_count": 0,
                    "failed_count": 0,
                    "total_count": 0
                })
        
        # Return aggregated results
        total_products = len(item_numbers)
        success_products = sum(1 for r in results if r.get("success"))
        failure_products = total_products - success_products
        overall_success = failure_products == 0
        
        return {
            "success": overall_success,
            "total_products": total_products,
            "success_products": success_products,
            "failure_products": failure_products,
            "total_uploaded_images": total_uploaded_images,
            "total_failed_images": total_failed_images,
            "total_images": total_images,
            "results": results,
            "message": f"Uploaded images for {success_products}/{total_products} products successfully ({total_uploaded_images} images uploaded)" + (f" ({failure_products} products failed)" if failure_products > 0 else "")
        }
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Exception while uploading multiple images to Rakuten: {error_msg}")
        add_log("error", "Exception while uploading multiple images to Rakuten", error_msg, "rakuten")
        
        return {
            "success": False,
            "error": error_msg,
            "message": f"Failed to upload images to Rakuten: {error_msg}"
        }


@app.delete("/api/product-management/{item_number}/images")
async def delete_product_image_endpoint(item_number: str, request: DeleteImageRequest):
    """
    Delete an image from the images field in product_management table.
    
    Request body:
    {
        "location": "/01306503/01306503_4.jpg"
    }
    """
    try:
        image_location = request.location
        if not image_location:
            return {
                "success": False,
                "error": "location is required",
                "message": "Please provide image location"
            }
        
        logger.info(f"Deleting image {image_location} from product {item_number}")
        
        success = delete_product_image(item_number, image_location)
        
        if success:
            add_log("info", f"Image deleted", f"Deleted image {image_location} from product {item_number}", "product_management")
            return {
                "success": True,
                "message": f"Image deleted successfully"
            }
        else:
            return {
                "success": False,
                "error": "Image not found or deletion failed",
                "message": f"Could not delete image {image_location} from product {item_number}"
            }
    except Exception as e:
        logger.error(f"Failed to delete image for product {item_number}: {e}")
        add_log("error", "Failed to delete image", str(e), "product_management")
        return {
            "success": False,
            "error": str(e),
            "message": f"Failed to delete image: {str(e)}"
        }


@app.post("/api/product-management/register-to-rakuten")
async def register_product_to_rakuten(request: RegisterToRakutenRequest):
    """Register a product from product_management table to Rakuten API."""
    try:
        item_number = request.item_number
        if not item_number:
            return {
                "success": False,
                "error": "item_number is required",
                "message": "Please provide item_number"
            }
        
        # Check rakuten_registration_status before processing
        from modules.db import get_product_management_by_item_number
        product_data = get_product_management_by_item_number(item_number)
        if not product_data:
            return {
                "success": False,
                "error": "Product not found in database",
                "message": f"Product {item_number} not found in database"
            }
        
        current_status = product_data.get("rakuten_registration_status")
        # Only process if status is NULL, empty, 'onsale', 'true', or 'false'
        if current_status not in [None, "", "onsale", "true", "false"]:
            logger.info(f"Skipping product {item_number}: rakuten_registration_status is '{current_status}' (only NULL, '', 'onsale', 'true', or 'false' are allowed)")
            return {
                "success": False,
                "error": f"Product registration status '{current_status}' does not allow registration",
                "message": f"登録ステータスが '{current_status}' のため、登録をスキップしました",
                "skipped": True
            }
        
        # Register product to Rakuten using the rakuten_product module
        result = register_product_from_product_management(item_number)
        
        if result.get("success"):
            logger.info(f"Successfully registered product {item_number} to Rakuten")
            add_log("success", f"Product {item_number} registered to Rakuten", result.get("message", ""), "rakuten")
            
            # Update registration status in database (success)
            from modules.db import update_rakuten_registration_status
            update_rakuten_registration_status(item_number, "true")
            
            return {
                "success": True,
                "message": result.get("message", "Product registered successfully"),
                "data": result.get("data")
            }
        else:
            error_msg = result.get("error", "Unknown error")
            error_details = format_error_message(result)
            
            # Log detailed error information
            logger.error(f"Failed to register product {item_number} to Rakuten: {error_msg}")
            if result.get("error_data"):
                logger.error(f"Rakuten API Error Data: {json.dumps(result.get('error_data'), indent=2, ensure_ascii=False)}")
            if result.get("error_text"):
                logger.error(f"Rakuten API Error Text: {result.get('error_text')}")
            if result.get("status_code"):
                logger.error(f"Rakuten API Status Code: {result.get('status_code')}")
            
            add_log("error", f"Failed to register product {item_number} to Rakuten", error_msg, "rakuten")
            
            # Update registration status in database (failed)
            from modules.db import update_rakuten_registration_status
            update_rakuten_registration_status(item_number, "false")
            
            return {
                "success": False,
                "error": error_msg,
                "error_details": error_details,
                "message": f"Failed to register product to Rakuten: {error_msg}"
            }
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Exception while registering product to Rakuten: {error_msg}")
        add_log("error", "Exception while registering product to Rakuten", error_msg, "rakuten")
        
        # Update registration status in database (failed)
        try:
            item_number = request.item_number
            if item_number:
                from modules.db import update_rakuten_registration_status
                update_rakuten_registration_status(item_number, "false")
        except Exception as db_error:
            logger.error(f"Failed to update registration status: {db_error}")
        
        return {
            "success": False,
            "error": error_msg,
            "message": f"Failed to register product to Rakuten: {error_msg}"
        }


@app.post("/api/product-management/check-registration-status")
async def check_product_registration_status(request: CheckRegistrationStatusRequest):
    """Check product registration status on Rakuten and update database accordingly."""
    try:
        item_number = request.item_number
        if not item_number:
            return {
                "success": False,
                "error": "item_number is required",
                "message": "Please provide item_number"
            }
        
        # Check and update registration status
        result = update_product_registration_status_from_rakuten(item_number)
        
        if result.get("success"):
            logger.info(f"Successfully checked registration status for product {item_number}: {result.get('status')}")
            add_log(
                "success" if result.get("status") == "registered" else "info",
                f"Checked registration status for product {item_number}",
                result.get("message", ""),
                "rakuten"
            )
            
            return {
                "success": True,
                "status": result.get("status"),
                "previous_status": result.get("previous_status"),
                "new_status": result.get("new_status"),
                "message": result.get("message", "Status checked successfully")
            }
        else:
            error_msg = result.get("error", "Unknown error")
            logger.error(f"Failed to check registration status for product {item_number}: {error_msg}")
            add_log("error", f"Failed to check registration status for product {item_number}", error_msg, "rakuten")
            
            return {
                "success": False,
                "error": error_msg,
                "status_code": result.get("status_code"),
                "error_data": result.get("error_data"),
                "message": f"Failed to check product status: {error_msg}"
            }
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Exception while checking registration status: {error_msg}")
        add_log("error", "Exception while checking registration status", error_msg, "rakuten")
        
        return {
            "success": False,
            "error": error_msg,
            "message": f"Failed to check product status: {error_msg}"
        }


@app.post("/api/product-management/check-multiple-registration-status")
async def check_multiple_products_registration_status(request: CheckMultipleRegistrationStatusRequest):
    """Check registration status for multiple products on Rakuten and update database accordingly."""
    try:
        item_numbers = request.item_numbers
        if not item_numbers or len(item_numbers) == 0:
            return {
                "success": False,
                "error": "item_numbers is required",
                "message": "Please provide at least one item_number"
            }
        
        # Check and update registration status for all products
        result = update_multiple_products_registration_status_from_rakuten(item_numbers)
        
        if result.get("success"):
            success_count = result.get("success_count", 0)
            error_count = result.get("error_count", 0)
            total = result.get("total", len(item_numbers))
            
            logger.info(f"Checked registration status for {total} products: {success_count} successful, {error_count} errors")
            add_log(
                "info",
                f"Checked registration status for {total} products",
                f"{success_count} successful, {error_count} errors",
                "rakuten"
            )
            
            return {
                "success": True,
                "total": total,
                "success_count": success_count,
                "error_count": error_count,
                "results": result.get("results", []),
                "message": f"Checked {total} products: {success_count} successful, {error_count} errors"
            }
        else:
            error_msg = result.get("error", "Unknown error")
            logger.error(f"Failed to check registration status for multiple products: {error_msg}")
            add_log("error", "Failed to check registration status for multiple products", error_msg, "rakuten")
            
            return {
                "success": False,
                "error": error_msg,
                "message": f"Failed to check product status: {error_msg}"
            }
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Exception while checking multiple products registration status: {error_msg}")
        add_log("error", "Exception while checking multiple products registration status", error_msg, "rakuten")
        
        return {
            "success": False,
            "error": error_msg,
            "message": f"Failed to check product status: {error_msg}"
        }


@app.post("/api/product-management/register-inventory-to-rakuten")
async def register_inventory_to_rakuten(request: RegisterInventoryToRakutenRequest):
    """Register inventory from product_management table to Rakuten API."""
    try:
        item_number = request.item_number
        if not item_number:
            return {
                "success": False,
                "error": "item_number is required",
                "message": "Please provide item_number"
            }
        
        # Register inventory to Rakuten using the rakuten_inventory module
        result = register_inventory_from_product_management(item_number)
        
        if result.get("success"):
            registered_count = result.get("registered_count", 0)
            total_count = result.get("total_count", 0)
            failed_count = result.get("failed_count", 0)
            
            if failed_count and failed_count > 0:
                logger.info(f"Partially registered inventory for product {item_number}: {registered_count}/{total_count} variants")
                add_log("warning", f"Partially registered inventory for product {item_number}", f"{registered_count}/{total_count} variants registered", "rakuten")
                
                # Update inventory_registration_status to true if at least one variant was registered
                if registered_count > 0:
                    from modules.db import update_product_registration_status
                    try:
                        update_product_registration_status(
                            item_number,
                            inventory_registration_status=True
                        )
                        logger.info(f"Updated inventory_registration_status to true for product {item_number} (partial success)")
                    except Exception as e:
                        logger.warning(f"Failed to update inventory_registration_status for product {item_number}: {e}")
                
                return {
                    "success": True,
                    "message": result.get("message", f"Registered inventory for {registered_count}/{total_count} variants"),
                    "registered_count": registered_count,
                    "failed_count": failed_count,
                    "total_count": total_count,
                    "results": result.get("results", []),
                    "errors": result.get("errors", [])
                }
            else:
                logger.info(f"Successfully registered inventory for product {item_number}: {registered_count} variants")
                add_log("success", f"Inventory registered for product {item_number}", result.get("message", ""), "rakuten")
                
                # Update inventory_registration_status to true on successful registration
                from modules.db import update_product_registration_status
                try:
                    update_product_registration_status(
                        item_number,
                        inventory_registration_status=True
                    )
                    logger.info(f"Updated inventory_registration_status to true for product {item_number}")
                except Exception as e:
                    logger.warning(f"Failed to update inventory_registration_status for product {item_number}: {e}")
                
                return {
                    "success": True,
                    "message": result.get("message", f"Successfully registered inventory for all {registered_count} variants"),
                    "registered_count": registered_count,
                    "total_count": total_count,
                    "results": result.get("results", [])
                }
        else:
            error_msg = result.get("error", "Unknown error")
            error_details = None
            if result.get("errors"):
                error_details = "\n".join([f"Variant {e.get('variant_id', 'unknown')}: {e.get('error', 'Unknown error')}" for e in result.get("errors", [])])
            
            logger.error(f"Failed to register inventory for product {item_number} to Rakuten: {error_msg}")
            add_log("error", f"Failed to register inventory for product {item_number} to Rakuten", error_msg, "rakuten")
            return {
                "success": False,
                "error": error_msg,
                "error_details": error_details,
                "message": f"Failed to register inventory to Rakuten: {error_msg}",
                "errors": result.get("errors", [])
            }
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Exception while registering inventory to Rakuten: {error_msg}")
        add_log("error", "Exception while registering inventory to Rakuten", error_msg, "rakuten")
        return {
            "success": False,
            "error": error_msg,
            "message": f"Failed to register inventory to Rakuten: {error_msg}"
        }


@app.post("/api/product-management/register-multiple-inventory-to-rakuten")
async def register_multiple_inventory_to_rakuten(request: RegisterMultipleInventoryToRakutenRequest):
    """Register inventory for multiple products from product_management table to Rakuten API sequentially."""
    try:
        item_numbers = request.item_numbers
        if not item_numbers or len(item_numbers) == 0:
            return {
                "success": False,
                "error": "item_numbers is required",
                "message": "Please provide at least one item_number"
            }
        
        from modules.db import update_product_registration_status
        
        results = []
        total_registered_count = 0
        total_failed_count = 0
        total_variants_count = 0
        
        # Process each product sequentially
        for idx, item_number in enumerate(item_numbers, 1):
            if not item_number:
                continue
                
            try:
                logger.info(f"Registering inventory for product {idx}/{len(item_numbers)}: {item_number}")
                
                # Register inventory to Rakuten using the rakuten_inventory module
                result = register_inventory_from_product_management(item_number)
                
                if result.get("success"):
                    registered_count = result.get("registered_count", 0)
                    total_count = result.get("total_count", 0)
                    failed_count = result.get("failed_count", 0)
                    
                    total_registered_count += registered_count
                    total_failed_count += failed_count
                    total_variants_count += total_count
                    
                    if failed_count and failed_count > 0:
                        logger.info(f"Partially registered inventory for product {item_number}: {registered_count}/{total_count} variants")
                        add_log("warning", f"Partially registered inventory for product {item_number}", f"{registered_count}/{total_count} variants registered", "rakuten")
                        
                        # Update inventory_registration_status to true if at least one variant was registered
                        if registered_count > 0:
                            try:
                                update_product_registration_status(
                                    item_number,
                                    inventory_registration_status=True
                                )
                                logger.info(f"Updated inventory_registration_status to true for product {item_number} (partial success)")
                            except Exception as e:
                                logger.warning(f"Failed to update inventory_registration_status for product {item_number}: {e}")
                    else:
                        logger.info(f"Successfully registered inventory for product {item_number}: {registered_count} variants")
                        add_log("success", f"Inventory registered for product {item_number}", result.get("message", ""), "rakuten")
                        
                        # Update inventory_registration_status to true on successful registration
                        try:
                            update_product_registration_status(
                                item_number,
                                inventory_registration_status=True
                            )
                            logger.info(f"Updated inventory_registration_status to true for product {item_number}")
                        except Exception as e:
                            logger.warning(f"Failed to update inventory_registration_status for product {item_number}: {e}")
                    
                    results.append({
                        "item_number": item_number,
                        "success": True,
                        "registered_count": registered_count,
                        "failed_count": failed_count,
                        "total_count": total_count,
                        "message": result.get("message", f"Registered inventory for {registered_count}/{total_count} variants"),
                        "results": result.get("results", []),
                        "errors": result.get("errors", [])
                    })
                else:
                    error_msg = result.get("error", "Unknown error")
                    error_details = None
                    if result.get("errors"):
                        error_details = "\n".join([f"Variant {e.get('variant_id', 'unknown')}: {e.get('error', 'Unknown error')}" for e in result.get("errors", [])])
                    
                    logger.error(f"Failed to register inventory for product {item_number} to Rakuten: {error_msg}")
                    add_log("error", f"Failed to register inventory for product {item_number} to Rakuten", error_msg, "rakuten")
                    
                    results.append({
                        "item_number": item_number,
                        "success": False,
                        "error": error_msg,
                        "error_details": error_details,
                        "message": f"Failed to register inventory to Rakuten: {error_msg}",
                        "errors": result.get("errors", [])
                    })
                    total_failed_count += 1
                    
            except Exception as e:
                error_msg = str(e)
                logger.error(f"Exception while registering inventory for product {item_number} to Rakuten: {error_msg}")
                add_log("error", f"Exception while registering inventory for product {item_number} to Rakuten", error_msg, "rakuten")
                
                results.append({
                    "item_number": item_number,
                    "success": False,
                    "error": error_msg,
                    "message": f"Failed to register inventory to Rakuten: {error_msg}"
                })
                total_failed_count += 1
        
        # Return aggregated results
        total_products = len(item_numbers)
        success_products = sum(1 for r in results if r.get("success"))
        failure_products = total_products - success_products
        overall_success = failure_products == 0
        
        return {
            "success": overall_success,
            "total_products": total_products,
            "success_products": success_products,
            "failure_products": failure_products,
            "total_registered_variants": total_registered_count,
            "total_failed_variants": total_failed_count,
            "total_variants": total_variants_count,
            "results": results,
            "message": f"Registered inventory for {success_products}/{total_products} products successfully ({total_registered_count} variants registered)" + (f" ({failure_products} products failed)" if failure_products > 0 else "")
        }
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Exception while registering multiple inventory to Rakuten: {error_msg}")
        add_log("error", "Exception while registering multiple inventory to Rakuten", error_msg, "rakuten")
        
        return {
            "success": False,
            "error": error_msg,
            "message": f"Failed to register inventory to Rakuten: {error_msg}"
        }


@app.delete("/api/product-management/{item_number}/delete-from-rakuten")
async def delete_product_from_rakuten(item_number: str):
    """Delete a product from Rakuten API using item_number from product_management table."""
    try:
        if not item_number:
            return {
                "success": False,
                "error": "item_number is required",
                "message": "Please provide item_number"
            }
        
        # Delete product from Rakuten using the rakuten_product module
        result = delete_product_from_product_management(item_number)
        
        if result.get("success"):
            logger.info(f"Successfully deleted product {item_number} from Rakuten")
            add_log("success", f"Product {item_number} deleted from Rakuten", result.get("message", ""), "rakuten")
            
            # Update registration status in database (set to unregistered/null to indicate deleted)
            from modules.db import update_rakuten_registration_status
            update_rakuten_registration_status(item_number, "unregistered")
            
            return {
                "success": True,
                "message": result.get("message", "商品を楽天市場から削除しました"),
                "data": result.get("data")
            }
        else:
            error_msg = result.get("error", "Unknown error")
            error_details = format_error_message(result)
            
            # Log detailed error information
            logger.error(f"Failed to delete product {item_number} from Rakuten: {error_msg}")
            if result.get("error_data"):
                logger.error(f"Rakuten API Error Data: {json.dumps(result.get('error_data'), indent=2, ensure_ascii=False)}")
            if result.get("error_text"):
                logger.error(f"Rakuten API Error Text: {result.get('error_text')}")
            if result.get("status_code"):
                logger.error(f"Rakuten API Status Code: {result.get('status_code')}")
            
            add_log("error", f"Failed to delete product {item_number} from Rakuten", error_msg, "rakuten")
            
            return {
                "success": False,
                "error": error_msg,
                "error_details": error_details,
                "message": f"Failed to delete product from Rakuten: {error_msg}"
            }
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Exception while deleting product from Rakuten: {error_msg}")
        add_log("error", "Exception while deleting product from Rakuten", error_msg, "rakuten")
        
        return {
            "success": False,
            "error": error_msg,
            "message": f"Failed to delete product from Rakuten: {error_msg}"
        }


@app.post("/api/product-management/delete-multiple-from-rakuten")
async def delete_multiple_products_from_rakuten(request: DeleteMultipleFromRakutenRequest):
    """Delete multiple products from Rakuten API sequentially."""
    try:
        item_numbers = request.item_numbers
        if not item_numbers or len(item_numbers) == 0:
            return {
                "success": False,
                "error": "item_numbers is required",
                "message": "Please provide at least one item_number"
            }
        
        from modules.db import update_rakuten_registration_status
        
        results = []
        success_count = 0
        failure_count = 0
        
        # Process each product sequentially
        for idx, item_number in enumerate(item_numbers, 1):
            if not item_number:
                continue
                
            try:
                logger.info(f"Deleting product {idx}/{len(item_numbers)} from Rakuten: {item_number}")
                
                # Delete product from Rakuten using the rakuten_product module
                result = delete_product_from_product_management(item_number)
                
                if result.get("success"):
                    logger.info(f"Successfully deleted product {item_number} from Rakuten")
                    add_log("success", f"Product {item_number} deleted from Rakuten", result.get("message", ""), "rakuten")
                    
                    # Update registration status in database (set to unregistered/null to indicate deleted)
                    update_rakuten_registration_status(item_number, "unregistered")
                    
                    results.append({
                        "item_number": item_number,
                        "success": True,
                        "message": result.get("message", "商品を楽天市場から削除しました")
                    })
                    success_count += 1
                else:
                    error_msg = result.get("error", "Unknown error")
                    error_details = format_error_message(result)
                    
                    # Log detailed error information
                    logger.error(f"Failed to delete product {item_number} from Rakuten: {error_msg}")
                    if result.get("error_data"):
                        logger.error(f"Rakuten API Error Data: {json.dumps(result.get('error_data'), indent=2, ensure_ascii=False)}")
                    if result.get("error_text"):
                        logger.error(f"Rakuten API Error Text: {result.get('error_text')}")
                    if result.get("status_code"):
                        logger.error(f"Rakuten API Status Code: {result.get('status_code')}")
                    
                    add_log("error", f"Failed to delete product {item_number} from Rakuten", error_msg, "rakuten")
                    
                    results.append({
                        "item_number": item_number,
                        "success": False,
                        "error": error_msg,
                        "error_details": error_details,
                        "message": f"削除に失敗しました: {error_msg}"
                    })
                    failure_count += 1
                    
            except Exception as e:
                error_msg = str(e)
                logger.error(f"Exception while deleting product {item_number} from Rakuten: {error_msg}")
                add_log("error", f"Exception while deleting product {item_number} from Rakuten", error_msg, "rakuten")
                
                results.append({
                    "item_number": item_number,
                    "success": False,
                    "error": error_msg,
                    "message": f"削除に失敗しました: {error_msg}"
                })
                failure_count += 1
        
        # Return aggregated results
        total_count = len(item_numbers)
        overall_success = failure_count == 0
        
        return {
            "success": overall_success,
            "total_count": total_count,
            "success_count": success_count,
            "failure_count": failure_count,
            "results": results,
            "message": f"{success_count}/{total_count}件の商品を楽天市場から削除しました" + (f"（{failure_count}件失敗）" if failure_count > 0 else "")
        }
            
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Exception while deleting multiple products from Rakuten: {error_msg}")
        add_log("error", "Exception while deleting multiple products from Rakuten", error_msg, "rakuten")
        
        return {
            "success": False,
            "error": error_msg,
            "message": f"削除に失敗しました: {error_msg}"
        }


@app.get("/api/product-management/export-csv")
async def export_product_management_csv():
    """Export product_management data as CSV file with UTF-8 BOM for Excel compatibility."""
    try:
        # Get all products (no pagination for export)
        items = get_product_management(limit=100000, offset=0, sort_by=None, sort_order=None)
        
        # Create CSV in memory with UTF-8 BOM
        output = io.StringIO()
        
        # Write UTF-8 BOM for Excel compatibility
        output.write('\ufeff')
        
        # Create CSV writer
        writer = csv.writer(output, quoting=csv.QUOTE_ALL, lineterminator='\n')
        
        # Write header
        writer.writerow([
            '商品番号',
            '商品名',
            '商品タグ',
            '商品説明文',
            '販売説明文',
            'ジャンルID',
            'タグ番号',
            'Rakumart URL',
            'Rakuten URL'
        ])
        
        # Write data rows
        for item in items:
            item_number = item.get('item_number', '')
            title = item.get('title', '')
            tagline = item.get('tagline', '')
            
            # Extract product_description (JSONB)
            product_description = ''
            product_desc_json = item.get('product_description')
            if product_desc_json:
                if isinstance(product_desc_json, dict):
                    # Combine pc and sp if both exist, otherwise use whichever is available
                    pc = product_desc_json.get('pc', '')
                    sp = product_desc_json.get('sp', '')
                    if pc and sp:
                        product_description = f"{pc}\n\n{sp}"
                    else:
                        product_description = pc or sp
                elif isinstance(product_desc_json, str):
                    try:
                        desc_obj = json.loads(product_desc_json)
                        pc = desc_obj.get('pc', '')
                        sp = desc_obj.get('sp', '')
                        if pc and sp:
                            product_description = f"{pc}\n\n{sp}"
                        else:
                            product_description = pc or sp
                    except:
                        product_description = product_desc_json
            
            sales_description = item.get('sales_description', '')
            genre_id = item.get('genre_id', '')
            
            # Extract tags (array)
            tags_str = ''
            tags = item.get('tags')
            if tags:
                if isinstance(tags, list):
                    tags_str = ','.join(str(t) for t in tags if t)
                elif isinstance(tags, str):
                    try:
                        tags_list = json.loads(tags)
                        if isinstance(tags_list, list):
                            tags_str = ','.join(str(t) for t in tags_list if t)
                    except:
                        tags_str = str(tags)
            
            src_url = item.get('src_url', '')
            rakuten_url = f"https://item.rakuten.co.jp/licel-store/{item_number}/" if item_number else ''
            
            writer.writerow([
                item_number,
                title,
                tagline,
                product_description,
                sales_description,
                genre_id,
                tags_str,
                src_url,
                rakuten_url
            ])
        
        # Get CSV content
        csv_content = output.getvalue()
        output.close()
        
        # Create response with UTF-8 BOM
        def generate():
            yield csv_content.encode('utf-8-sig')  # UTF-8 with BOM
        
        return StreamingResponse(
            generate(),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename=product_management_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            }
        )
        
    except Exception as e:
        logger.error(f"Failed to export CSV: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export CSV: {str(e)}")


@app.get("/api/product-management", response_model=ProductResponse)
async def list_product_management(
    limit: int = 50, 
    offset: int = 0, 
    sort_by: Optional[str] = None, 
    sort_order: Optional[str] = None
):
    """
    List products from product_management table with optional sorting.
    
    Args:
        limit: Maximum number of products to return
        offset: Number of products to skip
        sort_by: Field to sort by ('created_at' or 'rakuten_registered_at')
        sort_order: Sort order ('asc' or 'desc')
    """
    try:
        items = get_product_management(limit=limit, offset=offset, sort_by=sort_by, sort_order=sort_order)
        return ProductResponse(success=True, data=items, total=len(items), message=f"Retrieved {len(items)} items")
    except Exception as e:
        logger.error(f"Failed to list product_management: {e}")
        return ProductResponse(success=False, error=str(e), message="Failed to list product management")


@app.get("/api/stats", response_model=StatsResponse)
async def get_stats():
    """Basic counts for dashboard KPIs."""
    try:
        counts = get_counts()
        recent_products = get_recently_registered_products(limit=5)
        category_counts = get_category_registration_counts()
        counts["recent_products"] = recent_products
        counts["category_registration_counts"] = category_counts
        return StatsResponse(success=True, data=counts)
    except Exception as e:
        logger.error(f"Failed to get stats: {e}")
        return StatsResponse(success=False, error=str(e))


@app.get("/api/product-management/stats")
async def get_product_management_stats_endpoint():
    """Get detailed statistics for product_management table."""
    try:
        stats = get_product_management_stats()
        return {
            "success": True,
            "data": stats
        }
    except Exception as e:
        logger.error(f"Failed to get product management stats: {e}")
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/database/fix-schema", response_model=DatabaseResponse)
async def fix_database_schema():
    """
    Fix database schema issues
    """
    try:
        logger.info("Fixing database schema")
        fix_products_origin_schema()
        
        return DatabaseResponse(
            success=True,
            message="Database schema fixed successfully"
        )
        
    except Exception as e:
        logger.error(f"Failed to fix database schema: {e}")
        return DatabaseResponse(
            success=False,
            error=str(e),
            message="Failed to fix database schema"
        )

@app.post("/api/database/drop-removed-columns", response_model=DatabaseResponse)
async def drop_removed_columns():
    """
    Drop removed columns from products_origin table.
    This migration removes deprecated fields that are no longer used.
    """
    try:
        logger.info("Dropping removed columns from products_origin table")
        drop_removed_columns_from_products_origin()
        
        add_log(
            "success",
            "Removed columns dropped",
            "Successfully dropped removed columns from products_origin table",
            "database"
        )
        
        return DatabaseResponse(
            success=True,
            message="Removed columns dropped successfully from products_origin table"
        )
        
    except Exception as e:
        logger.error(f"Failed to drop removed columns: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        add_log(
            "error",
            "Failed to drop removed columns",
            f"Error: {str(e)}",
            "database"
        )
        return DatabaseResponse(
            success=False,
            error=str(e),
            message=f"Failed to drop removed columns: {str(e)}"
        )

@app.post("/api/product-management/import-csv")
async def import_product_management_csv(file: UploadFile = File(...)):
    """Import product_management data from CSV file."""
    try:
        # Read CSV file
        content = await file.read()
        
        # Decode with UTF-8 BOM support
        try:
            text = content.decode('utf-8-sig')  # Handles BOM
        except UnicodeDecodeError:
            try:
                text = content.decode('shift_jis')  # Fallback to Shift-JIS
            except UnicodeDecodeError:
                text = content.decode('utf-8', errors='ignore')
        
        # Parse CSV
        csv_reader = csv.DictReader(io.StringIO(text))
        
        updated_count = 0
        error_count = 0
        errors = []
        
        for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (header is row 1)
            try:
                item_number = row.get('商品番号', '').strip()
                if not item_number:
                    errors.append(f"行 {row_num}: 商品番号が空です")
                    error_count += 1
                    continue
                
                # Get existing product
                existing = get_product_management_by_item_number(item_number)
                if not existing:
                    errors.append(f"行 {row_num}: 商品番号 '{item_number}' が見つかりません")
                    error_count += 1
                    continue
                
                # Prepare update data
                update_data = {}
                
                # Update title if provided
                title = row.get('商品名', '').strip()
                if title:
                    update_data['title'] = title
                
                # Update tagline if provided
                tagline = row.get('商品タグ', '').strip()
                if tagline:
                    update_data['tagline'] = tagline
                
                # Update product_description if provided
                product_description = row.get('商品説明文', '').strip()
                if product_description:
                    # Parse existing product_description to preserve structure
                    existing_desc = existing.get('product_description')
                    if isinstance(existing_desc, dict):
                        # Update pc field, keep sp if exists
                        update_data['product_description'] = {
                            'pc': product_description,
                            'sp': existing_desc.get('sp', product_description)
                        }
                    elif isinstance(existing_desc, str):
                        try:
                            desc_obj = json.loads(existing_desc)
                            desc_obj['pc'] = product_description
                            if 'sp' not in desc_obj:
                                desc_obj['sp'] = product_description
                            update_data['product_description'] = desc_obj
                        except:
                            update_data['product_description'] = {
                                'pc': product_description,
                                'sp': product_description
                            }
                    else:
                        update_data['product_description'] = {
                            'pc': product_description,
                            'sp': product_description
                        }
                
                # Update sales_description if provided
                sales_description = row.get('販売説明文', '').strip()
                if sales_description:
                    update_data['sales_description'] = sales_description
                
                # Update genre_id if provided
                genre_id = row.get('ジャンルID', '').strip()
                if genre_id:
                    update_data['genre_id'] = genre_id
                
                # Update tags if provided
                tags_str = row.get('タグ番号', '').strip()
                if tags_str:
                    try:
                        # Parse comma-separated tags
                        tags_list = [int(t.strip()) for t in tags_str.split(',') if t.strip().isdigit()]
                        if tags_list:
                            update_data['tags'] = tags_list
                    except Exception as e:
                        logger.warning(f"Failed to parse tags for {item_number}: {e}")
                
                # Update src_url if provided
                src_url = row.get('Rakumart URL', '').strip()
                if src_url:
                    update_data['src_url'] = src_url
                
                # Perform update if we have data to update
                if update_data:
                    update_product_management_settings(
                        item_number,
                        title=update_data.get('title'),
                        tagline=update_data.get('tagline'),
                        product_description=update_data.get('product_description'),
                        sales_description=update_data.get('sales_description'),
                        genre_id=update_data.get('genre_id'),
                        tags=update_data.get('tags'),
                        src_url=update_data.get('src_url')
                    )
                    updated_count += 1
                else:
                    errors.append(f"行 {row_num}: 更新するデータがありません")
                    
            except Exception as e:
                error_msg = f"行 {row_num}: {str(e)}"
                errors.append(error_msg)
                error_count += 1
                logger.error(f"Error processing CSV row {row_num}: {e}", exc_info=True)
        
        message = f"{updated_count}件の商品を更新しました"
        if error_count > 0:
            message += f"（{error_count}件エラー）"
        
        return {
            "success": True,
            "updated_count": updated_count,
            "error_count": error_count,
            "errors": errors[:50],  # Limit to first 50 errors
            "message": message
        }
        
    except Exception as e:
        logger.error(f"Failed to import CSV: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to import CSV: {str(e)}")


@app.get("/api/product-management", response_model=ProductResponse)
async def list_product_management(limit: int = 50, offset: int = 0):
    try:
        items = get_product_management(limit=limit, offset=offset)
        return ProductResponse(success=True, data=items, total=len(items), message=f"Retrieved {len(items)} items")
    except Exception as e:
        logger.error(f"Failed to list product_management: {e}")
        return ProductResponse(success=False, error=str(e), message="Failed to list product management")


@app.post("/api/database/reset-product-management", response_model=DatabaseResponse)
async def reset_product_management():
    """
    Reset the product_management table by dropping and recreating it with the new schema.
    This will delete all existing data in the product_management table.
    WARNING: This operation is irreversible!
    """
    try:
        logger.warning("Resetting product_management table - all data will be lost!")
        reset_product_management_table()
        
        add_log(
            "success",
            "Product management table reset",
            "Successfully reset product_management table with new schema",
            "database"
        )
        
        return DatabaseResponse(
            success=True,
            message="Product management table reset successfully with new schema matching Rakuten API structure"
        )
        
    except Exception as e:
        logger.error(f"Failed to reset product_management table: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        add_log(
            "error",
            "Failed to reset product management table",
            f"Error: {str(e)}",
            "database"
        )
        return DatabaseResponse(
            success=False,
            error=str(e),
            message=f"Failed to reset product_management table: {str(e)}"
        )

# Settings endpoints
@app.get("/api/settings", response_model=SettingsResponse)
async def get_settings():
    """
    Get current settings (loads pricing settings from database)
    """
    try:
        # Reload settings to get latest from database
        load_settings()
        
        return SettingsResponse(
            success=True,
            settings=SettingsRequest(**settings_data)
        )
    except Exception as e:
        logger.error(f"Failed to get settings: {e}")
        add_log("error", "Failed to get settings", str(e), "settings")
        return SettingsResponse(
            success=False,
            error=str(e)
        )

@app.post("/api/settings", response_model=SettingsResponse)
async def update_settings(request: SettingsRequest):
    """
    Update settings (pricing settings to database, others to file)
    """
    try:
        global settings_data
        shipping_costs = request.domestic_shipping_costs.dict()
        regular_shipping_cost = shipping_costs.get("regular", request.domestic_shipping_cost)
        
        # Save pricing settings to database
        try:
            ensure_settings_table()
            save_pricing_settings(
                exchange_rate=request.exchange_rate,
                profit_margin_percent=request.profit_margin_percent,
                sales_commission_percent=request.sales_commission_percent,
                currency=request.currency,
                domestic_shipping_cost=regular_shipping_cost,
                domestic_shipping_costs=shipping_costs,
                international_shipping_rate=request.international_shipping_rate,
                customs_duty_rate=request.customs_duty_rate,
            )
            logger.info("✅ Saved pricing settings to database")
        except Exception as e:
            logger.warning(f"⚠️  Failed to save pricing settings to database: {e}. Saving to file instead.")
        
        # Save all settings to file (for backward compatibility and non-pricing settings)
        settings_dict = request.dict()
        settings_dict["domestic_shipping_cost"] = regular_shipping_cost
        settings_dict["domestic_shipping_costs"] = shipping_costs
        settings_data = settings_dict
        save_settings()
        
        add_log("info", "Settings updated successfully", f"Auto refresh: {request.auto_refresh}, Logging: {request.logging_enabled}", "settings")
        
        # Restart auto-refresh task if settings changed
        await start_auto_refresh_task()
        
        return SettingsResponse(
            success=True,
            settings=request
        )
    except Exception as e:
        logger.error(f"Failed to update settings: {e}")
        add_log("error", "Failed to update settings", str(e), "settings")
        return SettingsResponse(
            success=False,
            error=str(e)
        )

# Primary category management endpoints
@app.get("/api/settings/primary-categories", response_model=PrimaryCategoryListResponse)
async def get_primary_categories_endpoint():
    """
    Return all registered primary categories.
    """
    try:
        categories = list_primary_categories()
        return PrimaryCategoryListResponse(
            success=True,
            categories=[PrimaryCategoryRecord(**category) for category in categories],
        )
    except Exception as e:
        logger.error(f"Failed to load primary categories: {e}")
        add_log("error", "Failed to load primary categories", str(e), "settings")
        return PrimaryCategoryListResponse(success=False, error=str(e))


@app.post("/api/settings/primary-categories", response_model=PrimaryCategoryMutationResponse)
async def create_primary_category_endpoint(request: PrimaryCategoryCreateRequest):
    """
    Create a new primary category.
    """
    try:
        if not request.category_name.strip():
            raise HTTPException(status_code=400, detail="Category name is required.")

        category = create_primary_category(
            category_name=request.category_name,
            default_category_ids=request.default_category_ids
        )
        add_log("success", "Primary category registered", f"Primary category '{category['category_name']}' added", "settings")
        return PrimaryCategoryMutationResponse(success=True, category=PrimaryCategoryRecord(**category))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to create primary category: {e}")
        add_log("error", "Failed to create primary category", str(e), "settings")
        return PrimaryCategoryMutationResponse(success=False, error=str(e))


@app.put("/api/settings/primary-categories/{category_id}", response_model=PrimaryCategoryMutationResponse)
async def update_primary_category_endpoint(category_id: int, request: PrimaryCategoryUpdateRequest):
    """
    Update an existing primary category.
    """
    try:
        update_kwargs = {}
        if request.category_name is not None:
            if not request.category_name.strip():
                raise HTTPException(status_code=400, detail="Category name cannot be empty.")
            update_kwargs["category_name"] = request.category_name
        if request.default_category_ids is not None:
            update_kwargs["default_category_ids"] = request.default_category_ids
        
        if not update_kwargs:
            raise HTTPException(status_code=400, detail="At least one field must be provided for update.")

        category = update_primary_category(category_id, **update_kwargs)
        if not category:
            raise HTTPException(status_code=404, detail="Primary category not found.")

        add_log("success", "Primary category updated", f"Primary category '{category['category_name']}' updated", "settings")
        return PrimaryCategoryMutationResponse(success=True, category=PrimaryCategoryRecord(**category))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update primary category: {e}")
        add_log("error", "Failed to update primary category", str(e), "settings")
        return PrimaryCategoryMutationResponse(success=False, error=str(e))


@app.delete("/api/settings/primary-categories/{category_id}", response_model=PrimaryCategoryMutationResponse)
async def delete_primary_category_endpoint(category_id: int):
    """
    Delete a primary category.
    """
    try:
        deleted = delete_primary_category(category_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Primary category not found.")

        add_log("success", "Primary category deleted", f"Primary category ID {category_id} removed", "settings")
        return PrimaryCategoryMutationResponse(success=True, category=None)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete primary category: {e}")
        add_log("error", "Failed to delete primary category", str(e), "settings")
        return PrimaryCategoryMutationResponse(success=False, error=str(e))


@app.get("/api/settings/primary-categories/export")
async def export_primary_categories_endpoint():
    """
    Export all primary categories to XLSX file.
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="XLSX export not available. Please install openpyxl.")
    
    try:
        categories = list_primary_categories()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Primary Categories"
        
        # Headers
        headers = ["ID", "カテゴリ名", "デフォルトカテゴリID (JSON)", "作成日時", "更新日時"]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for cat in categories:
            default_ids_str = json.dumps(cat.get("default_category_ids", []), ensure_ascii=False)
            row = [
                cat.get("id"),
                cat.get("category_name", ""),
                default_ids_str,
                cat.get("created_at").isoformat() if cat.get("created_at") else "",
                cat.get("updated_at").isoformat() if cat.get("updated_at") else "",
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"primary_categories_export_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        logger.error(f"Failed to export primary categories: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export primary categories: {str(e)}")


@app.post("/api/settings/primary-categories/import")
async def import_primary_categories_endpoint(file: UploadFile = File(...)):
    """
    Import primary categories from XLSX file.
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="XLSX import not available. Please install openpyxl.")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only XLSX files are supported.")
    
    try:
        # Read file content
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        # Read headers
        headers = [cell.value for cell in ws[1]]
        
        # Expected headers mapping
        header_map = {
            "ID": "id",
            "カテゴリ名": "category_name",
            "デフォルトカテゴリID (JSON)": "default_category_ids",
        }
        
        # Find column indices
        col_indices = {}
        for idx, header in enumerate(headers):
            if header in header_map:
                col_indices[header_map[header]] = idx
        
        if "category_name" not in col_indices:
            raise HTTPException(status_code=400, detail="Required column 'カテゴリ名' not found.")
        
        # Process rows
        imported_count = 0
        updated_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Skip empty rows
                if not row[col_indices.get("category_name")]:
                    continue
                
                category_name = str(row[col_indices.get("category_name")]).strip()
                if not category_name:
                    continue
                
                # Parse JSON fields
                default_category_ids = []
                if "default_category_ids" in col_indices and row[col_indices["default_category_ids"]]:
                    try:
                        default_category_ids = json.loads(str(row[col_indices["default_category_ids"]]))
                    except:
                        pass
                
                # Check if category exists (by ID)
                category_id = None
                if "id" in col_indices and row[col_indices["id"]]:
                    try:
                        category_id = int(row[col_indices["id"]])
                    except:
                        pass
                
                if category_id:
                    # Update existing category
                    update_kwargs = {
                        "category_name": category_name,
                    }
                    if "default_category_ids" in col_indices:
                        update_kwargs["default_category_ids"] = default_category_ids
                    
                    updated = update_primary_category(category_id, **update_kwargs)
                    if updated:
                        updated_count += 1
                    else:
                        errors.append(f"Row {row_idx}: Primary category ID {category_id} not found for update")
                else:
                    # Create new category
                    created = create_primary_category(
                        category_name=category_name,
                        default_category_ids=default_category_ids
                    )
                    if created:
                        imported_count += 1
                    else:
                        errors.append(f"Row {row_idx}: Failed to create primary category")
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                logger.error(f"Error processing row {row_idx}: {e}", exc_info=True)
        
        result_message = f"Imported {imported_count} new primary categories, updated {updated_count} primary categories."
        if errors:
            result_message += f" {len(errors)} errors occurred."
        
        add_log("success", "Primary categories imported", result_message, "settings")
        
        return JSONResponse({
            "success": True,
            "imported": imported_count,
            "updated": updated_count,
            "errors": errors[:10],  # Limit errors to first 10
            "message": result_message
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to import primary categories: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to import primary categories: {str(e)}")

# Category management endpoints
@app.get("/api/settings/categories", response_model=CategoryListResponse)
async def get_categories_endpoint():
    """
    Return all registered category management entries.
    """
    try:
        categories = list_categories()
        return CategoryListResponse(
            success=True,
            categories=[CategoryRecord(**category) for category in categories],
        )
    except Exception as e:
        logger.error(f"Failed to load categories: {e}")
        add_log("error", "Failed to load category list", str(e), "settings")
        return CategoryListResponse(success=False, error=str(e))


@app.post("/api/settings/categories", response_model=CategoryMutationResponse)
async def create_category_endpoint(request: CategoryCreateRequest):
    """
    Create a new category entry.
    """
    try:
        if not request.category_ids:
            raise HTTPException(status_code=400, detail="At least one category ID is required.")

        name = request.category_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Category name is required.")

        # Convert attributes (Pydantic models) to plain dicts for DB layer
        attributes_payload = (
            [attr.dict() for attr in request.attributes]
            if request.attributes is not None
            else None
        )

        category = create_category_entry(
            category_name=name,
            category_ids=request.category_ids,
            rakuten_category_ids=request.rakuten_category_ids,
            genre_id=request.genre_id,
            primary_category_id=request.primary_category_id,
            weight=request.weight,
            length=request.length,
            width=request.width,
            height=request.height,
            size_option=request.size_option,
            size=request.size,
            attributes=attributes_payload,
        )
        add_log("success", "Category registered", f"Category '{request.category_name}' added", "settings")
        return CategoryMutationResponse(success=True, category=CategoryRecord(**category))
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error("Failed to create category", exc_info=True)
        add_log("error", "Failed to create category", repr(e), "settings")
        return CategoryMutationResponse(success=False, error=str(e))


@app.put("/api/settings/categories/{category_id}", response_model=CategoryMutationResponse)
async def update_category_endpoint(category_id: int, request: CategoryUpdateRequest):
    """
    Update an existing category entry.
    """
    try:
        update_kwargs = {}
        if request.category_name is not None:
            trimmed_name = request.category_name.strip()
            if not trimmed_name:
                raise HTTPException(status_code=400, detail="Category name cannot be empty.")
            update_kwargs["category_name"] = trimmed_name
        if request.category_ids is not None:
            if not request.category_ids:
                raise HTTPException(status_code=400, detail="At least one category ID is required.")
            update_kwargs["category_ids"] = request.category_ids
        if request.rakuten_category_ids is not None:
            update_kwargs["rakuten_category_ids"] = request.rakuten_category_ids
        if request.genre_id is not None or "genre_id" in request.__fields_set__:
            update_kwargs["genre_id"] = request.genre_id
        if "primary_category_id" in request.__fields_set__:
            update_kwargs["primary_category_id"] = request.primary_category_id
        if request.weight is not None or "weight" in request.__fields_set__:
            update_kwargs["weight"] = request.weight
        if request.length is not None or "length" in request.__fields_set__:
            update_kwargs["length"] = request.length
        if request.width is not None or "width" in request.__fields_set__:
            update_kwargs["width"] = request.width
        if request.height is not None or "height" in request.__fields_set__:
            update_kwargs["height"] = request.height
        if request.size_option is not None or "size_option" in request.__fields_set__:
            update_kwargs["size_option"] = request.size_option
        if request.size is not None or "size" in request.__fields_set__:
            update_kwargs["size"] = request.size
        if request.attributes is not None or "attributes" in request.__fields_set__:
            update_kwargs["attributes"] = (
                [attr.dict() for attr in request.attributes]
                if request.attributes is not None
                else []
            )

        if not update_kwargs:
            raise HTTPException(status_code=400, detail="No category fields were provided for update.")

        category = update_category_entry(category_id, **update_kwargs)
        if not category:
            raise HTTPException(status_code=404, detail="Category not found.")

        add_log("success", "Category updated", f"Category '{category.get('category_name')}' updated", "settings")
        return CategoryMutationResponse(success=True, category=CategoryRecord(**category))
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error("Failed to update category", exc_info=True)
        add_log("error", "Failed to update category", repr(e), "settings")
        return CategoryMutationResponse(success=False, error=str(e))


@app.delete("/api/settings/categories/{category_id}", response_model=CategoryMutationResponse)
async def delete_category_endpoint(category_id: int):
    """
    Delete a category entry.
    """
    try:
        deleted = delete_category_entry(category_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Category not found.")

        add_log("success", "Category deleted", f"Category ID {category_id} removed", "settings")
        return CategoryMutationResponse(success=True, category=None)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete category: {e}")
        add_log("error", "Failed to delete category", str(e), "settings")
        return CategoryMutationResponse(success=False, error=str(e))


@app.get("/api/settings/categories/export")
async def export_categories_endpoint():
    """
    Export all categories to XLSX file.
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="XLSX export not available. Please install openpyxl.")
    
    try:
        categories = list_categories()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Categories"
        
        # Headers
        headers = [
            "ID", "カテゴリ名", "メインカテゴリID", "メインカテゴリ名", 
            "カテゴリID (JSON)", "楽天カテゴリID (JSON)", "ジャンルID",
            "重量 (kg)", "長さ (cm)", "幅 (cm)", "高さ (cm)", "サイズオプション", "サイズ (cm)",
            "属性 (JSON)", "作成日時", "更新日時"
        ]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for cat in categories:
            # Convert JSON fields to strings for Excel
            category_ids_str = json.dumps(cat.get("category_ids", []), ensure_ascii=False)
            rakuten_ids_str = json.dumps(cat.get("rakuten_category_ids", []), ensure_ascii=False)
            attributes_str = json.dumps(cat.get("attributes", []), ensure_ascii=False)
            
            row = [
                cat.get("id"),
                cat.get("category_name", ""),
                cat.get("primary_category_id"),
                cat.get("primary_category_name", ""),
                category_ids_str,
                rakuten_ids_str,
                cat.get("genre_id", ""),
                cat.get("weight"),
                cat.get("length"),
                cat.get("width"),
                cat.get("height"),
                cat.get("size_option", ""),
                cat.get("size"),
                attributes_str,
                cat.get("created_at").isoformat() if cat.get("created_at") else "",
                cat.get("updated_at").isoformat() if cat.get("updated_at") else "",
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"categories_export_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        logger.error(f"Failed to export categories: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export categories: {str(e)}")


# Risk Product Settings (JSON file-based)
RISK_PRODUCTS_JSON_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "risk_products.json")

def ensure_risk_products_file():
    """Ensure the risk products JSON file exists with default structure"""
    os.makedirs(os.path.dirname(RISK_PRODUCTS_JSON_PATH), exist_ok=True)
    if not os.path.exists(RISK_PRODUCTS_JSON_PATH):
        default_data = {
            "high_risk": {
                "keywords": [],
                "category_ids": []
            },
            "low_risk": {
                "keywords": [],
                "category_ids": []
            }
        }
        with open(RISK_PRODUCTS_JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(default_data, f, ensure_ascii=False, indent=2)
        logger.info(f"Created default risk products file at {RISK_PRODUCTS_JSON_PATH}")

def load_risk_products() -> dict:
    """Load risk products data from JSON file"""
    ensure_risk_products_file()
    try:
        with open(RISK_PRODUCTS_JSON_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Failed to load risk products: {e}")
        return {
            "high_risk": {"keywords": [], "category_ids": []},
            "low_risk": {"keywords": [], "category_ids": []}
        }

def save_risk_products(data: dict):
    """Save risk products data to JSON file"""
    ensure_risk_products_file()
    try:
        with open(RISK_PRODUCTS_JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"Saved risk products to {RISK_PRODUCTS_JSON_PATH}")
    except Exception as e:
        logger.error(f"Failed to save risk products: {e}")
        raise

@app.get("/api/settings/risk-products")
async def get_risk_products():
    """Get risk products settings (keywords and category IDs)"""
    try:
        data = load_risk_products()
        return {
            "success": True,
            "data": data
        }
    except Exception as e:
        logger.error(f"Failed to get risk products: {e}")
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/settings/risk-products")
async def update_risk_products(request: dict = Body(...)):
    """Update risk products settings (keywords and category IDs)"""
    try:
        # Validate structure
        if "high_risk" not in request or "low_risk" not in request:
            return {
                "success": False,
                "error": "Invalid structure. Must contain 'high_risk' and 'low_risk'"
            }
        
        # Validate each risk level
        for risk_level in ["high_risk", "low_risk"]:
            if risk_level not in request:
                continue
            risk_data = request[risk_level]
            if "keywords" not in risk_data or "category_ids" not in risk_data:
                return {
                    "success": False,
                    "error": f"Invalid structure for {risk_level}. Must contain 'keywords' and 'category_ids'"
                }
            # Ensure arrays
            if not isinstance(risk_data["keywords"], list):
                risk_data["keywords"] = []
            if not isinstance(risk_data["category_ids"], list):
                risk_data["category_ids"] = []
        
        # Save to file
        save_risk_products(request)
        
        return {
            "success": True,
            "data": request,
            "message": "Risk products settings updated successfully"
        }
    except Exception as e:
        logger.error(f"Failed to update risk products: {e}")
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/settings/categories/import")
async def import_categories(file: UploadFile = File(...)):
    """
    Import categories from CSV file
    """
    try:
        content = await file.read()
        csv_content = content.decode('utf-8-sig')  # Handle BOM
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        
        imported_count = 0
        errors = []
        
        for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (1 is header)
            try:
                category_name = row.get('category_name', '').strip()
                category_ids_str = row.get('category_ids', '').strip()
                genre_id = row.get('genre_id', '').strip() or None
                
                if not category_name:
                    errors.append(f"Row {row_num}: category_name is required")
                    continue
                
                # Parse category_ids
                category_ids = []
                if category_ids_str:
                    # Split by comma and clean up
                    category_ids = [cid.strip() for cid in category_ids_str.split(',') if cid.strip()]
                
                # Create category
                result = create_category_entry(
                    category_name=category_name,
                    category_ids=category_ids,
                    genre_id=genre_id
                )
                
                if result:
                    imported_count += 1
                else:
                    errors.append(f"Row {row_num}: Failed to create category")
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        if errors:
            logger.warning(f"Import completed with {len(errors)} errors")
        
        return {
            "success": True,
            "imported_count": imported_count,
            "error_count": len(errors),
            "errors": errors[:50],  # Limit to first 50 errors
            "message": f"Imported {imported_count} categories"
        }
    except Exception as e:
        logger.error(f"Failed to import categories: {e}")
        return {
            "success": False,
            "error": str(e),
            "message": "Failed to import categories"
        }


# Risk Product Settings (JSON file-based)
RISK_PRODUCTS_JSON_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "risk_products.json")

def ensure_risk_products_file():
    """Ensure the risk products JSON file exists with default structure"""
    os.makedirs(os.path.dirname(RISK_PRODUCTS_JSON_PATH), exist_ok=True)
    if not os.path.exists(RISK_PRODUCTS_JSON_PATH):
        default_data = {
            "high_risk": {
                "keywords": [],
                "category_ids": []
            },
            "low_risk": {
                "keywords": [],
                "category_ids": []
            }
        }
        with open(RISK_PRODUCTS_JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(default_data, f, ensure_ascii=False, indent=2)
        logger.info(f"Created default risk products file at {RISK_PRODUCTS_JSON_PATH}")

def load_risk_products() -> dict:
    """Load risk products data from JSON file"""
    ensure_risk_products_file()
    try:
        with open(RISK_PRODUCTS_JSON_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Failed to load risk products: {e}")
        return {
            "high_risk": {"keywords": [], "category_ids": []},
            "low_risk": {"keywords": [], "category_ids": []}
        }

def save_risk_products(data: dict):
    """Save risk products data to JSON file"""
    ensure_risk_products_file()
    try:
        with open(RISK_PRODUCTS_JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"Saved risk products to {RISK_PRODUCTS_JSON_PATH}")
    except Exception as e:
        logger.error(f"Failed to save risk products: {e}")
        raise

@app.get("/api/settings/risk-products")
async def get_risk_products():
    """Get risk products settings (keywords and category IDs)"""
    try:
        data = load_risk_products()
        return {
            "success": True,
            "data": data
        }
    except Exception as e:
        logger.error(f"Failed to get risk products: {e}")
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/settings/risk-products")
async def update_risk_products(request: dict = Body(...)):
    """Update risk products settings (keywords and category IDs)"""
    try:
        # Validate structure
        if "high_risk" not in request or "low_risk" not in request:
            return {
                "success": False,
                "error": "Invalid structure. Must contain 'high_risk' and 'low_risk'"
            }
        
        # Validate each risk level
        for risk_level in ["high_risk", "low_risk"]:
            if risk_level not in request:
                continue
            risk_data = request[risk_level]
            if "keywords" not in risk_data or "category_ids" not in risk_data:
                return {
                    "success": False,
                    "error": f"Invalid structure for {risk_level}. Must contain 'keywords' and 'category_ids'"
                }
            # Ensure arrays
            if not isinstance(risk_data["keywords"], list):
                risk_data["keywords"] = []
            if not isinstance(risk_data["category_ids"], list):
                risk_data["category_ids"] = []
        
        # Save to file
        save_risk_products(request)
        
        return {
            "success": True,
            "data": request,
            "message": "Risk products settings updated successfully"
        }
    except Exception as e:
        logger.error(f"Failed to update risk products: {e}")
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/settings/categories/import")
async def import_categories_endpoint(file: UploadFile = File(...)):
    """
    Import categories from XLSX file.
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="XLSX import not available. Please install openpyxl.")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only XLSX files are supported.")
    
    try:
        # Read file content
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        # Read headers
        headers = [cell.value for cell in ws[1]]
        
        # Expected headers mapping
        header_map = {
            "ID": "id",
            "カテゴリ名": "category_name",
            "メインカテゴリID": "primary_category_id",
            "カテゴリID (JSON)": "category_ids",
            "楽天カテゴリID (JSON)": "rakuten_category_ids",
            "ジャンルID": "genre_id",
            "重量 (kg)": "weight",
            "長さ (cm)": "length",
            "幅 (cm)": "width",
            "高さ (cm)": "height",
            "サイズオプション": "size_option",
            "サイズ (cm)": "size",
            "属性 (JSON)": "attributes",
        }
        
        # Find column indices
        col_indices = {}
        for idx, header in enumerate(headers):
            if header in header_map:
                col_indices[header_map[header]] = idx
        
        if "category_name" not in col_indices:
            raise HTTPException(status_code=400, detail="Required column 'カテゴリ名' not found.")
        if "category_ids" not in col_indices:
            raise HTTPException(status_code=400, detail="Required column 'カテゴリID (JSON)' not found.")
        
        # Process rows
        imported_count = 0
        updated_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Skip empty rows
                if not row[col_indices.get("category_name")]:
                    continue
                
                category_name = str(row[col_indices.get("category_name")]).strip()
                if not category_name:
                    continue
                
                # Parse JSON fields
                category_ids = []
                if "category_ids" in col_indices and row[col_indices["category_ids"]]:
                    try:
                        category_ids = json.loads(str(row[col_indices["category_ids"]]))
                    except:
                        errors.append(f"Row {row_idx}: Invalid category_ids JSON")
                        continue
                
                if not category_ids:
                    errors.append(f"Row {row_idx}: category_ids is required")
                    continue
                
                rakuten_category_ids = None
                if "rakuten_category_ids" in col_indices and row[col_indices["rakuten_category_ids"]]:
                    try:
                        rakuten_category_ids = json.loads(str(row[col_indices["rakuten_category_ids"]]))
                    except:
                        pass
                
                genre_id = None
                if "genre_id" in col_indices and row[col_indices["genre_id"]]:
                    genre_id = str(row[col_indices["genre_id"]]).strip() or None
                
                primary_category_id = None
                if "primary_category_id" in col_indices and row[col_indices["primary_category_id"]]:
                    try:
                        primary_category_id = int(row[col_indices["primary_category_id"]]) if row[col_indices["primary_category_id"]] else None
                    except:
                        pass
                
                # Parse numeric fields
                weight = None
                if "weight" in col_indices and row[col_indices["weight"]]:
                    try:
                        weight = float(row[col_indices["weight"]])
                    except:
                        pass
                
                length = None
                if "length" in col_indices and row[col_indices["length"]]:
                    try:
                        length = float(row[col_indices["length"]])
                    except:
                        pass
                
                width = None
                if "width" in col_indices and row[col_indices["width"]]:
                    try:
                        width = float(row[col_indices["width"]])
                    except:
                        pass
                
                height = None
                if "height" in col_indices and row[col_indices["height"]]:
                    try:
                        height = float(row[col_indices["height"]])
                    except:
                        pass
                
                size = None
                if "size" in col_indices and row[col_indices["size"]]:
                    try:
                        size = float(row[col_indices["size"]])
                    except:
                        pass
                
                size_option = None
                if "size_option" in col_indices and row[col_indices["size_option"]]:
                    size_option = str(row[col_indices["size_option"]]).strip() or None
                
                attributes = None
                if "attributes" in col_indices and row[col_indices["attributes"]]:
                    try:
                        attributes = json.loads(str(row[col_indices["attributes"]]))
                    except:
                        pass
                
                # Check if category exists (by ID or by name)
                category_id = None
                if "id" in col_indices and row[col_indices["id"]]:
                    try:
                        category_id = int(row[col_indices["id"]])
                    except:
                        pass
                
                if category_id:
                    # Update existing category
                    update_kwargs = {
                        "category_name": category_name,
                        "category_ids": category_ids,
                    }
                    if rakuten_category_ids is not None:
                        update_kwargs["rakuten_category_ids"] = rakuten_category_ids
                    if genre_id is not None or "genre_id" in col_indices:
                        update_kwargs["genre_id"] = genre_id
                    if primary_category_id is not None or "primary_category_id" in col_indices:
                        update_kwargs["primary_category_id"] = primary_category_id
                    if weight is not None or "weight" in col_indices:
                        update_kwargs["weight"] = weight
                    if length is not None or "length" in col_indices:
                        update_kwargs["length"] = length
                    if width is not None or "width" in col_indices:
                        update_kwargs["width"] = width
                    if height is not None or "height" in col_indices:
                        update_kwargs["height"] = height
                    if size is not None or "size" in col_indices:
                        update_kwargs["size"] = size
                    if size_option is not None or "size_option" in col_indices:
                        update_kwargs["size_option"] = size_option
                    if attributes is not None or "attributes" in col_indices:
                        update_kwargs["attributes"] = attributes
                    
                    updated = update_category_entry(category_id, **update_kwargs)
                    if updated:
                        updated_count += 1
                    else:
                        errors.append(f"Row {row_idx}: Category ID {category_id} not found for update")
                else:
                    # Create new category
                    created = create_category_entry(
                        category_name=category_name,
                        category_ids=category_ids,
                        rakuten_category_ids=rakuten_category_ids,
                        genre_id=genre_id,
                        primary_category_id=primary_category_id,
                        weight=weight,
                        length=length,
                        width=width,
                        height=height,
                        size_option=size_option,
                        size=size,
                        attributes=attributes,
                    )
                    if created:
                        imported_count += 1
                    else:
                        errors.append(f"Row {row_idx}: Failed to create category")
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                logger.error(f"Error processing row {row_idx}: {e}", exc_info=True)
        
        result_message = f"Imported {imported_count} new categories, updated {updated_count} categories."
        if errors:
            result_message += f" {len(errors)} errors occurred."
        
        add_log("success", "Categories imported", result_message, "settings")
        
        return JSONResponse({
            "success": True,
            "imported": imported_count,
            "updated": updated_count,
            "errors": errors[:10],  # Limit errors to first 10
            "message": result_message
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to import categories: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to import categories: {str(e)}")

# Logs endpoints
@app.get("/api/logs", response_model=LogsResponse)
async def get_logs(limit: int = 100):
    """
    Get system logs
    """
    try:
        logs_to_return = logs_data[:limit] if limit > 0 else logs_data
        
        return LogsResponse(
            success=True,
            logs=[LogEntry(**log) for log in logs_to_return],
            total_count=len(logs_data)
        )
    except Exception as e:
        logger.error(f"Failed to get logs: {e}")
        add_log("error", "Failed to get logs", str(e), "logs")
        return LogsResponse(
            success=False,
            error=str(e)
        )

@app.delete("/api/logs")
async def clear_logs():
    """
    Clear all logs
    """
    try:
        global logs_data
        logs_data = []
        save_logs()
        
        add_log("info", "Logs cleared successfully", "All logs have been cleared", "logs")
        
        return {"success": True, "message": "Logs cleared successfully"}
    except Exception as e:
        logger.error(f"Failed to clear logs: {e}")
        add_log("error", "Failed to clear logs", str(e), "logs")
        return {"success": False, "error": str(e)}

# Auto-refresh keyword management endpoints
@app.get("/api/auto-refresh/keywords")
async def get_refresh_keywords_endpoint():
    """
    Get the list of keywords for automatic refresh
    """
    try:
        keywords = get_refresh_keywords()
        return {"success": True, "keywords": keywords}
    except Exception as e:
        logger.error(f"Failed to get refresh keywords: {e}")
        add_log("error", "Failed to get refresh keywords", str(e), "auto-refresh")
        return {"success": False, "error": str(e)}

@app.post("/api/auto-refresh/keywords")
async def add_refresh_keyword_endpoint(keyword: str):
    """
    Add a keyword to the automatic refresh list
    """
    try:
        add_refresh_keyword(keyword)
        return {"success": True, "message": f"Keyword '{keyword}' added to auto-refresh"}
    except Exception as e:
        logger.error(f"Failed to add refresh keyword: {e}")
        add_log("error", "Failed to add refresh keyword", str(e), "auto-refresh")
        return {"success": False, "error": str(e)}

@app.delete("/api/auto-refresh/keywords/{keyword}")
async def remove_refresh_keyword_endpoint(keyword: str):
    """
    Remove a keyword from the automatic refresh list
    """
    try:
        remove_refresh_keyword(keyword)
        return {"success": True, "message": f"Keyword '{keyword}' removed from auto-refresh"}
    except Exception as e:
        logger.error(f"Failed to remove refresh keyword: {e}")
        add_log("error", "Failed to remove refresh keyword", str(e), "auto-refresh")
        return {"success": False, "error": str(e)}

@app.post("/api/auto-refresh/trigger")
async def trigger_manual_refresh():
    """
    Manually trigger an automatic refresh cycle
    """
    try:
        add_log("info", "Manual refresh triggered", "User requested immediate refresh", "auto-refresh")
        await perform_automatic_refresh()
        return {"success": True, "message": "Manual refresh completed"}
    except Exception as e:
        logger.error(f"Failed to trigger manual refresh: {e}")
        add_log("error", "Manual refresh failed", str(e), "auto-refresh")
        return {"success": False, "error": str(e)}

if __name__ == "__main__":
    import uvicorn
    import os
    
    # Configure uvicorn logging to suppress asyncio errors
    log_config = {
        "version": 1,
        "disable_existing_loggers": False,
        "formatters": {
            "default": {
                "format": "%(asctime)s - %(name)s - %(levelname)s - %(message)s",
            },
        },
        "handlers": {
            "default": {
                "formatter": "default",
                "class": "logging.StreamHandler",
                "stream": "ext://sys.stdout",
            },
        },
        "loggers": {
            "asyncio": {
                "handlers": ["default"],
                "level": "CRITICAL",
                "propagate": False,
            },
            "asyncio.proactor_events": {
                "handlers": ["default"],
                "level": "CRITICAL",
                "propagate": False,
            },
            "asyncio.windows_events": {
                "handlers": ["default"],
                "level": "CRITICAL",
                "propagate": False,
            },
        },
        "root": {
            "level": "INFO",
            "handlers": ["default"],
        },
    }
    
    # Allow port to be configured via environment variable, default to 8000
    port = int(os.getenv("API_PORT", "8000"))
    uvicorn.run(app, host="0.0.0.0", port=port, log_config=log_config)
